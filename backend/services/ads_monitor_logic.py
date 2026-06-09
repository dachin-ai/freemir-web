"""Ads Monitor — daily TikTok product campaign file (file ads) breakdown."""

from __future__ import annotations

import base64
import io
from datetime import datetime, timezone
from typing import Iterable

import numpy as np
import pandas as pd
from sqlalchemy.orm import Session

from models import (
    AdsMonitorDailyRecord,
    AdsMonitorDiscoveredAccount,
    AdsMonitorInternalCreator,
)

BUCKET_PRODUCT_CARD = "product_card"
BUCKET_INTERNAL = "internal_creator"
BUCKET_EXTERNAL = "external_creator"

# Excel columns D, G, K, N (0-based) for standard TikTok creative export.
_FALLBACK_COL_INDEX = {
    "Creative type": 3,
    "TikTok account": 6,
    "Cost": 10,
    "Gross revenue": 13,
}

_COL_ALIASES = {
    "Creative type": ("creative type", "creative_type"),
    "TikTok account": ("tiktok account", "account name", "account"),
    "Cost": ("cost", "spend"),
    "Gross revenue": ("gross revenue", "revenue", "gmv"),
}


def _normalize_account(value) -> str:
    return str(value or "").strip()


def _is_product_card_account(account: str) -> bool:
    return account in {"", "-"}


def _normalize_creative_type(value) -> str:
    return str(value or "").strip().lower()


def _is_product_card_creative(creative_type: str) -> bool:
    return _normalize_creative_type(creative_type) == "product card"


def _is_video_creative(creative_type: str) -> bool:
    return _normalize_creative_type(creative_type) == "video"


def _map_columns_by_header(headers: list) -> dict[str, str] | None:
    """Map source column label → canonical name. None if any required col missing."""
    lower_to_src: dict[str, str] = {}
    for h in headers:
        key = str(h or "").strip().lower()
        if key:
            lower_to_src[key] = h

    out: dict[str, str] = {}
    for canonical, aliases in _COL_ALIASES.items():
        found = None
        for alias in aliases:
            if alias in lower_to_src:
                found = lower_to_src[alias]
                break
        if not found:
            return None
        out[canonical] = found
    return out


def _excel_number_to_str(val) -> str:
    """Text representation for Excel Find/Replace '.' → ',' workflow."""
    if val is None:
        return ""
    if isinstance(val, (float, np.floating)):
        if pd.isna(val) or np.isinf(val):
            return ""
        text = format(float(val), ".6f").rstrip("0").rstrip(".")
        return text if text else "0"
    if isinstance(val, (int, np.integer)):
        return str(int(val))
    return str(val).strip()


def _parse_excel_id_decimal(val) -> float:
    """
    Match Indonesian Excel: Find/Replace '.' → ',' on dot-decimal cells, then sum.
    Also accepts values already in ID format (8.560.060,818).
    """
    text = _excel_number_to_str(val)
    if not text or text.lower() in {"nan", "n/a", "-", "none", "#n/a"}:
        return 0.0
    text = text.replace(" ", "")
    try:
        if "," in text:
            return float(text.replace(".", "").replace(",", "."))
        if "." in text:
            text = text.replace(".", ",")
            return float(text.replace(".", "").replace(",", "."))
        return float(text)
    except ValueError:
        return 0.0


def _to_float_series(s: pd.Series) -> pd.Series:
    return s.map(_parse_excel_id_decimal)


def _normalize_ads_monitor_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df["Creative type"] = df["Creative type"].map(_normalize_creative_type)
    df["TikTok account"] = df["TikTok account"].fillna("-").astype(str).str.strip()
    df.loc[df["TikTok account"] == "", "TikTok account"] = "-"
    return df


def _filter_meaningful_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    Keep rows with Cost > 0 OR Gross revenue > 0 (matches manual Excel pivot scope).
    Drops only rows where both metrics are zero — vectorized, no row cap.
    """
    rows_in_file = len(df)
    if df.empty:
        return df, rows_in_file

    df = df.copy()
    df["Cost"] = _to_float_series(df["Cost"])
    df["Gross revenue"] = _to_float_series(df["Gross revenue"])
    df = df.loc[(df["Cost"] != 0) | (df["Gross revenue"] != 0)]
    if df.empty:
        return df.iloc[0:0].copy(), rows_in_file

    df = (
        df.sort_values(["Cost", "Gross revenue"], ascending=[False, False], kind="mergesort")
        .pipe(_normalize_ads_monitor_frame)
        .reset_index(drop=True)
    )
    return df, rows_in_file


def _parse_data_date(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        raise ValueError("Data date is required (YYYY-MM-DD).")
    ts = pd.to_datetime(raw, format="%Y-%m-%d", errors="coerce")
    if ts is None or pd.isna(ts):
        raise ValueError("Invalid data date. Use YYYY-MM-DD.")
    return pd.Timestamp(ts).strftime("%Y-%m-%d")


def _read_excel_openpyxl_meaningful_rows(file_bytes: bytes, sheet_name: str) -> tuple[pd.DataFrame, int]:
    """Fallback reader — stream rows, keep Cost>0 or Gross>0 only."""
    from openpyxl import load_workbook

    creatives: list[str] = []
    accounts: list[str] = []
    costs: list[float] = []
    grosses: list[float] = []
    rows_seen = 0

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for row in ws.iter_rows(
            min_row=2, max_row=250_000, min_col=1, max_col=14, values_only=True
        ):
            if not row:
                continue
            rows_seen += 1
            cost = _parse_excel_id_decimal(row[10] if len(row) > 10 else 0)
            gross = _parse_excel_id_decimal(row[13] if len(row) > 13 else 0)
            if cost == 0.0 and gross == 0.0:
                continue
            creatives.append(_normalize_creative_type(row[3] if len(row) > 3 else ""))
            acc = str(row[6] if len(row) > 6 else "-").strip() or "-"
            accounts.append(acc)
            costs.append(cost)
            grosses.append(gross)
    finally:
        wb.close()

    if not accounts:
        return pd.DataFrame(columns=list(_FALLBACK_COL_INDEX.keys())), rows_seen

    return pd.DataFrame(
        {
            "Creative type": creatives,
            "TikTok account": accounts,
            "Cost": costs,
            "Gross revenue": grosses,
        }
    ), rows_seen


def _read_excel_ads_monitor(file_bytes: bytes) -> tuple[pd.DataFrame, int]:
    """Read 4 columns via calamine (fast); openpyxl heap fallback if unavailable."""
    sheet = "Data"
    bio = io.BytesIO(file_bytes)

    try:
        header_df = pd.read_excel(bio, sheet_name=sheet, nrows=0, engine="calamine")
    except (ValueError, KeyError):
        bio.seek(0)
        header_df = pd.read_excel(bio, nrows=0, engine="calamine")
        sheet = 0

    headers = list(header_df.columns)
    col_map = _map_columns_by_header(headers)
    bio.seek(0)

    read_kwargs: dict = {"sheet_name": sheet, "engine": "calamine"}
    if col_map:
        read_kwargs["usecols"] = list(col_map.values())
        read_kwargs["dtype"] = {
            col_map["Cost"]: str,
            col_map["Gross revenue"]: str,
        }
        df = pd.read_excel(bio, **read_kwargs)
        df = df.rename(columns={src: canon for canon, src in col_map.items()})
    else:
        read_kwargs["usecols"] = list(_FALLBACK_COL_INDEX.values())
        df = pd.read_excel(bio, **read_kwargs)
        df.columns = list(_FALLBACK_COL_INDEX.keys())

    return _filter_meaningful_rows(df)


def _read_ads_monitor_dataframe(file_bytes: bytes, filename: str) -> tuple[pd.DataFrame, int]:
    """
    Load Creative type (D), TikTok account (G), Cost (K), Gross revenue (N).
    Keeps rows with Cost or Gross revenue (excludes both-zero noise rows).
    """
    if filename.lower().endswith(".csv"):
        bio = io.BytesIO(file_bytes)
        header_df = pd.read_csv(bio, nrows=0)
        headers = list(header_df.columns)
        col_map = _map_columns_by_header(headers)
        bio.seek(0)
        if col_map:
            df = pd.read_csv(bio, usecols=list(col_map.values()))
            df = df.rename(columns={src: canon for canon, src in col_map.items()})
        else:
            bio.seek(0)
            df = pd.read_csv(bio, usecols=list(_FALLBACK_COL_INDEX.values()))
            df.columns = list(_FALLBACK_COL_INDEX.keys())
        return _filter_meaningful_rows(df)

    try:
        return _read_excel_ads_monitor(file_bytes)
    except Exception:
        raw, rows_in_file = _read_excel_openpyxl_meaningful_rows(file_bytes, "Data")
        raw = _normalize_ads_monitor_frame(raw)
        return raw.reset_index(drop=True), rows_in_file


def _bucket_metrics_vectorized(df: pd.DataFrame, internal_lower: set[str]) -> dict[str, dict]:
    creative = df["Creative type"]
    is_product_card = creative.map(_is_product_card_creative)
    is_video = creative.map(_is_video_creative)

    acc_lower = df["TikTok account"].str.lower()
    is_internal = is_video & acc_lower.isin(internal_lower)
    is_external = is_video & ~is_internal

    cost = df["Cost"].to_numpy(dtype=np.float64, copy=False)
    gross = df["Gross revenue"].to_numpy(dtype=np.float64, copy=False)

    def pack(mask: np.ndarray) -> dict:
        c = float(cost[mask].sum())
        g = float(gross[mask].sum())
        roi = (g / c) if c > 0 else 0.0
        return {
            "cost": round(c, 3),
            "grossRevenue": round(g, 3),
            "roi": round(roi, 4),
            "rowCount": int(mask.sum()),
        }

    return {
        BUCKET_PRODUCT_CARD: pack(is_product_card.to_numpy()),
        BUCKET_INTERNAL: pack(is_internal.to_numpy()),
        BUCKET_EXTERNAL: pack(is_external.to_numpy()),
    }


def _internal_creator_breakdown(
    df: pd.DataFrame,
    internal_accounts: list[str],
    internal_lower: set[str],
) -> list[dict]:
    """Per-account Cost / Gross / ROI for configured internal creators (Video rows only)."""
    if not internal_accounts:
        return []

    is_video = df["Creative type"].map(_is_video_creative)
    acc_lower = df["TikTok account"].str.lower()
    internal_video = df.loc[is_video & acc_lower.isin(internal_lower)]

    by_key: dict[str, dict] = {}
    if not internal_video.empty:
        grouped = internal_video.groupby("TikTok account", as_index=False).agg(
            Cost=("Cost", "sum"),
            Gross_revenue=("Gross revenue", "sum"),
            rowCount=("Cost", "size"),
        )
        for _, row in grouped.iterrows():
            account = _normalize_account(row["TikTok account"])
            cost = float(row["Cost"])
            gross = float(row["Gross_revenue"])
            roi = (gross / cost) if cost > 0 else 0.0
            by_key[account.lower()] = {
                "account": account,
                "cost": round(cost, 3),
                "grossRevenue": round(gross, 3),
                "roi": round(roi, 4),
                "rowCount": int(row["rowCount"]),
            }

    out: list[dict] = []
    for account in internal_accounts:
        entry = by_key.get(account.lower())
        if entry:
            out.append(entry)
        else:
            out.append({
                "account": account,
                "cost": 0.0,
                "grossRevenue": 0.0,
                "roi": 0.0,
                "rowCount": 0,
            })

    out.sort(key=lambda item: (-item["cost"], item["account"].lower()))
    return out


def list_internal_creators(db: Session) -> list[str]:
    rows = (
        db.query(AdsMonitorInternalCreator.account_name)
        .order_by(AdsMonitorInternalCreator.account_name.asc())
        .all()
    )
    return [r[0] for r in rows if r[0]]


def save_internal_creators(db: Session, accounts: Iterable[str]) -> list[str]:
    cleaned = sorted({
        _normalize_account(a)
        for a in accounts
        if _normalize_account(a) and not _is_product_card_account(_normalize_account(a))
    })
    db.query(AdsMonitorInternalCreator).delete()
    if cleaned:
        db.bulk_save_objects([AdsMonitorInternalCreator(account_name=name) for name in cleaned])
    db.commit()
    return cleaned


def list_discovered_accounts(db: Session, *, query: str = "", limit: int = 500) -> list[str]:
    q = db.query(AdsMonitorDiscoveredAccount.account_name)
    needle = (query or "").strip().lower()
    if needle:
        q = q.filter(AdsMonitorDiscoveredAccount.account_name.ilike(f"%{needle}%"))
    rows = q.order_by(AdsMonitorDiscoveredAccount.account_name.asc()).limit(max(1, min(limit, 2000))).all()
    return [r[0] for r in rows if r[0]]


def _register_new_discovered_accounts(db: Session, accounts: Iterable[str]) -> int:
    """
    Insert only accounts not yet in DB (single IN lookup + bulk insert).
    Skips last_seen updates for existing rows — keeps analyze fast.
    """
    cleaned = sorted({
        _normalize_account(a)
        for a in accounts
        if _normalize_account(a) and not _is_product_card_account(_normalize_account(a))
    })
    if not cleaned:
        return 0

    existing = {
        r[0]
        for r in db.query(AdsMonitorDiscoveredAccount.account_name)
        .filter(AdsMonitorDiscoveredAccount.account_name.in_(cleaned))
        .all()
    }
    new_accounts = [name for name in cleaned if name not in existing]
    if not new_accounts:
        return 0

    now = datetime.now(timezone.utc)
    db.bulk_save_objects([
        AdsMonitorDiscoveredAccount(
            account_name=name,
            first_seen_at=now,
            last_seen_at=now,
        )
        for name in new_accounts
    ])
    db.commit()
    return len(new_accounts)


def get_tiktok_stores() -> list[dict]:
    """TikTok store codes from Admin Base > Store_Info."""
    from services.store_info_logic import get_tiktok_stores as _load_stores

    return _load_stores()


def get_tiktok_stores_payload(*, force_refresh: bool = False) -> dict:
    from services.store_info_logic import get_tiktok_stores_payload as _load_payload

    return _load_payload(force_refresh=force_refresh)


def _validate_store_code(store_code: str) -> str:
    code = (store_code or "").strip()
    if not code:
        raise ValueError("Store code is required.")
    allowed = {s["code"] for s in get_tiktok_stores()}
    if allowed and code not in allowed:
        raise ValueError(f"Unknown store code: {code}")
    return code


def _daily_record_to_day_payload(record: AdsMonitorDailyRecord) -> dict:
    return {
        "productCard": {
            "cost": float(record.product_card_cost or 0),
            "gmv": float(record.product_card_gmv or 0),
        },
        "internalCreator": {
            "cost": float(record.internal_cost or 0),
            "gmv": float(record.internal_gmv or 0),
        },
        "externalCreator": {
            "cost": float(record.external_cost or 0),
            "gmv": float(record.external_gmv or 0),
        },
    }


def save_manual_daily_record(
    db: Session,
    *,
    store_code: str,
    data_date: str,
    product_card_cost: float = 0,
    product_card_gmv: float = 0,
    internal_cost: float = 0,
    internal_gmv: float = 0,
    external_cost: float = 0,
    external_gmv: float = 0,
) -> dict:
    """Upsert daily snapshot from manual Cost/GMV entry."""
    data_date_str = _parse_data_date(data_date)
    store = _validate_store_code(store_code)
    buckets = {
        BUCKET_PRODUCT_CARD: {
            "cost": float(product_card_cost or 0),
            "grossRevenue": float(product_card_gmv or 0),
        },
        BUCKET_INTERNAL: {
            "cost": float(internal_cost or 0),
            "grossRevenue": float(internal_gmv or 0),
        },
        BUCKET_EXTERNAL: {
            "cost": float(external_cost or 0),
            "grossRevenue": float(external_gmv or 0),
        },
    }
    save_daily_record(
        db,
        store_code=store,
        data_date=data_date_str,
        buckets=buckets,
        filename="manual-import",
    )
    return {"storeCode": store, "dataDate": data_date_str}


def save_bulk_manual_records(
    db: Session,
    *,
    store_code: str,
    records: list[dict],
) -> dict:
    """Upsert many daily snapshots from import file rows (single commit)."""
    if not records:
        raise ValueError("No records to import.")
    store = _validate_store_code(store_code)

    parsed: list[dict] = []
    dates: list[str] = []
    for rec in records:
        data_date = rec.get("data_date") or rec.get("dataDate")
        if not data_date:
            continue
        data_date_str = _parse_data_date(str(data_date))
        pc = rec.get("product_card") or {}
        inh = rec.get("inhouse") or {}
        ext = rec.get("external") or {}
        parsed.append({
            "data_date": data_date_str,
            "product_card_cost": float(pc.get("cost") or 0),
            "product_card_gmv": float(pc.get("gmv") or 0),
            "internal_cost": float(inh.get("cost") or 0),
            "internal_gmv": float(inh.get("gmv") or 0),
            "external_cost": float(ext.get("cost") or 0),
            "external_gmv": float(ext.get("gmv") or 0),
        })
        dates.append(data_date_str)

    if not parsed:
        raise ValueError("No records to import.")

    date_keys = [row["data_date"] for row in parsed]
    existing = {
        row.data_date: row
        for row in db.query(AdsMonitorDailyRecord)
        .filter(
            AdsMonitorDailyRecord.store_code == store,
            AdsMonitorDailyRecord.data_date.in_(date_keys),
        )
        .all()
    }

    for row in parsed:
        record = existing.get(row["data_date"])
        if record is None:
            record = AdsMonitorDailyRecord(store_code=store, data_date=row["data_date"])
            db.add(record)
        record.product_card_cost = row["product_card_cost"]
        record.product_card_gmv = row["product_card_gmv"]
        record.internal_cost = row["internal_cost"]
        record.internal_gmv = row["internal_gmv"]
        record.external_cost = row["external_cost"]
        record.external_gmv = row["external_gmv"]
        record.source_filename = "manual-import"

    db.commit()
    return {
        "storeCode": store,
        "saved": len(parsed),
        "dates": sorted(set(dates)),
    }


def save_daily_record(
    db: Session,
    *,
    store_code: str,
    data_date: str,
    buckets: dict,
    filename: str = "",
) -> None:
    """Upsert daily Cost/GMV per segment for one store."""
    data_date_str = _parse_data_date(data_date)
    store = _validate_store_code(store_code)
    pc = buckets.get(BUCKET_PRODUCT_CARD, {})
    internal = buckets.get(BUCKET_INTERNAL, {})
    external = buckets.get(BUCKET_EXTERNAL, {})

    record = (
        db.query(AdsMonitorDailyRecord)
        .filter(
            AdsMonitorDailyRecord.store_code == store,
            AdsMonitorDailyRecord.data_date == data_date_str,
        )
        .first()
    )
    if record is None:
        record = AdsMonitorDailyRecord(store_code=store, data_date=data_date_str)
        db.add(record)

    record.product_card_cost = float(pc.get("cost") or 0)
    record.product_card_gmv = float(pc.get("grossRevenue") or 0)
    record.internal_cost = float(internal.get("cost") or 0)
    record.internal_gmv = float(internal.get("grossRevenue") or 0)
    record.external_cost = float(external.get("cost") or 0)
    record.external_gmv = float(external.get("grossRevenue") or 0)
    record.source_filename = filename or ""
    db.commit()


def get_monthly_report(db: Session, *, store_code: str, year: int, month: int) -> dict:
    """Daily snapshots for one store in a calendar month."""
    import calendar

    store = _validate_store_code(store_code)
    year = int(year)
    month = int(month)
    if month < 1 or month > 12:
        raise ValueError("Month must be between 1 and 12.")

    name_map = {s["code"]: s["name"] for s in get_tiktok_stores()}
    prefix = f"{year:04d}-{month:02d}-"
    rows = (
        db.query(AdsMonitorDailyRecord)
        .filter(
            AdsMonitorDailyRecord.store_code == store,
            AdsMonitorDailyRecord.data_date.like(f"{prefix}%"),
        )
        .order_by(AdsMonitorDailyRecord.data_date.asc())
        .all()
    )
    by_date = {row.data_date: _daily_record_to_day_payload(row) for row in rows}

    _, last_day = calendar.monthrange(year, month)
    dates: list[dict] = []
    for day in range(1, last_day + 1):
        iso = f"{year:04d}-{month:02d}-{day:02d}"
        ts = pd.Timestamp(iso)
        dates.append({
            "date": iso,
            "label": f"{month}/{day}/{year}",
            "weekday": ts.strftime("%A"),
            "hasData": iso in by_date,
            "segments": by_date.get(iso),
        })

    return {
        "storeCode": store,
        "storeName": name_map.get(store, store),
        "year": year,
        "month": month,
        "dates": dates,
        "filledDays": len(by_date),
    }


def delete_daily_record(db: Session, *, store_code: str, data_date: str) -> dict:
    """Remove one saved daily snapshot."""
    data_date_str = _parse_data_date(data_date)
    store = _validate_store_code(store_code)
    deleted = (
        db.query(AdsMonitorDailyRecord)
        .filter(
            AdsMonitorDailyRecord.store_code == store,
            AdsMonitorDailyRecord.data_date == data_date_str,
        )
        .delete(synchronize_session=False)
    )
    db.commit()
    return {"storeCode": store, "dataDate": data_date_str, "deleted": int(deleted or 0)}


def delete_month_records(db: Session, *, store_code: str, year: int, month: int) -> dict:
    """Remove all saved daily snapshots for one store in a calendar month."""
    store = _validate_store_code(store_code)
    year = int(year)
    month = int(month)
    if month < 1 or month > 12:
        raise ValueError("Month must be between 1 and 12.")
    prefix = f"{year:04d}-{month:02d}-"
    deleted = (
        db.query(AdsMonitorDailyRecord)
        .filter(
            AdsMonitorDailyRecord.store_code == store,
            AdsMonitorDailyRecord.data_date.like(f"{prefix}%"),
        )
        .delete(synchronize_session=False)
    )
    db.commit()
    return {
        "storeCode": store,
        "year": year,
        "month": month,
        "deleted": int(deleted or 0),
    }


def analyze_ads_file(
    db: Session,
    *,
    file_bytes: bytes,
    filename: str,
    data_date: str,
    store_code: str,
) -> dict:
    """Parse file ads — Product Card (creative type), Internal/External (video + kolom G)."""
    internal_accounts = list_internal_creators(db)
    internal_lower = {a.lower() for a in internal_accounts}
    data_date_str = _parse_data_date(data_date)

    df, rows_in_file = _read_ads_monitor_dataframe(file_bytes, filename)
    if df.empty:
        raise ValueError("No rows found in file.")

    video_mask = df["Creative type"].map(_is_video_creative)
    accounts_in_file = sorted({
        _normalize_account(acc)
        for acc in df.loc[video_mask, "TikTok account"].unique()
        if not _is_product_card_account(_normalize_account(acc))
    })
    new_accounts_count = _register_new_discovered_accounts(db, accounts_in_file)

    buckets = _bucket_metrics_vectorized(df, internal_lower)
    internal_breakdown = _internal_creator_breakdown(df, internal_accounts, internal_lower)
    total_cost = sum(b["cost"] for b in buckets.values())
    total_gross = sum(b["grossRevenue"] for b in buckets.values())
    total_roi = (total_gross / total_cost) if total_cost > 0 else 0.0

    store = _validate_store_code(store_code)
    save_daily_record(
        db,
        store_code=store,
        data_date=data_date_str,
        buckets=buckets,
        filename=filename,
    )

    return {
        "filename": filename,
        "storeCode": store,
        "dataDate": data_date_str,
        "rowCount": int(len(df)),
        "rowsInFile": int(rows_in_file),
        "accountsInFile": accounts_in_file,
        "newAccountsCount": new_accounts_count,
        "internalCreators": internal_accounts,
        "internalCreatorBreakdown": internal_breakdown,
        "buckets": buckets,
        "total": {
            "cost": round(total_cost, 3),
            "grossRevenue": round(total_gross, 3),
            "roi": round(total_roi, 4),
            "rowCount": int(len(df)),
        },
    }


def analyze_ads_file_b64(
    db: Session,
    *,
    filename: str,
    content_b64: str,
    data_date: str,
    store_code: str,
) -> dict:
    raw = base64.b64decode(content_b64)
    return analyze_ads_file(
        db,
        file_bytes=raw,
        filename=filename,
        data_date=data_date,
        store_code=store_code,
    )
