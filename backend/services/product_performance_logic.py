"""
Product Performance Cleaner — processing logic.

Shopee column mapping (0-indexed, from 商品表现明细 sheet):
  0  → PID
  1  → Product Picture
  2  → Product Name
  3  → Unit (Item Sold)
  4  → GMV
  6  → Impression (PV)
  8  → Clicks (Search clicks)
  9  → CTR
  14 → CO (Conversion rate)
"""

import io
import time
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, timedelta
import pandas as pd
import openpyxl
from sqlalchemy.orm import Session
from models import PidStoreMap, ProductPerformance

# ─────────────────────────────────────────────────────────────────
# Simple TTL cache for expensive Google Sheets calls
# ─────────────────────────────────────────────────────────────────
_CACHE: dict = {}
_CACHE_TTL = 300  # seconds (5 minutes)


def _cache_get(key: str):
    entry = _CACHE.get(key)
    if entry and time.time() - entry["ts"] < _CACHE_TTL:
        return entry["val"]
    return None


def _cache_set(key: str, val):
    _CACHE[key] = {"val": val, "ts": time.time()}
    return val

# ── Patch openpyxl to tolerate non-standard frozen-pane values in xlsx files ──
try:
    from openpyxl.descriptors.base import Set as _OpenpyxlSet
    _orig_set = _OpenpyxlSet.__set__
    def _lenient_set(self, instance, value):
        try:
            _orig_set(self, instance, value)
        except ValueError:
            pass
    _OpenpyxlSet.__set__ = _lenient_set
except Exception:
    pass

# ─────────────────────────────────────────────────────────────────
# WEEK UTILITIES
# ─────────────────────────────────────────────────────────────────

WEEK0_START = date(2025, 12, 22)  # Monday of Week 0


def get_week_label(week_num: int) -> str:
    start = WEEK0_START + timedelta(weeks=week_num)
    end = start + timedelta(days=6)
    return f"Week {week_num} ({start.strftime('%d-%b-%Y')} – {end.strftime('%d-%b-%Y')})"


def get_week_dates(week_num: int):
    start = WEEK0_START + timedelta(weeks=week_num)
    end = start + timedelta(days=6)
    return start.isoformat(), end.isoformat()


def get_all_weeks() -> list[dict]:
    """Return list of week options from Week 0 up to current week + 2."""
    today = date.today()
    days_since = (today - WEEK0_START).days
    max_week = max(0, days_since // 7 + 2)
    end_2026 = date(2026, 12, 31)
    max_week_2026 = max(0, (end_2026 - WEEK0_START).days // 7)
    max_week = max(max_week, max_week_2026)
    weeks = []
    for i in range(max_week + 1):
        start = WEEK0_START + timedelta(weeks=i)
        end = start + timedelta(days=6)
        weeks.append({
            "value": i,
            "label": f"Week {i}  ({start.strftime('%d-%b-%Y')} – {end.strftime('%d-%b-%Y')})",
            "start": start.isoformat(),
            "end": end.isoformat(),
        })
    return weeks


def get_store_mapping(db: Session) -> dict:
    """Return dict of pid → store (first store found for that PID)."""
    rows = db.query(PidStoreMap).all()
    result = {}
    for r in rows:
        if r.pid and r.pid not in result:
            result[r.pid] = r.store
    return result


def get_pid_to_store_map(db: Session) -> dict:
    rows = db.query(PidStoreMap).all()
    result = {}
    for r in rows:
        if r.pid and r.pid not in result:
            result[r.pid] = r.store
    return result


def get_pid_to_sku_map(db: Session) -> dict:
    rows = db.query(PidStoreMap).all()
    result = {}
    for r in rows:
        if r.pid and r.pid not in result:
            result[r.pid] = r.sku
    return result


def get_store_pid_to_sku_map(db: Session) -> dict:
    rows = db.query(PidStoreMap).all()
    result = {}
    for r in rows:
        if r.store and r.pid:
            result[(r.store, r.pid)] = r.sku
    return result


# Admin Base spreadsheet — worksheet "Store_Info" (gid=1046982230):
#   Column C → store code, Column E → full display name (1-based A=0 index below).
_STORE_INFO_COL_CODE = 2  # C
_STORE_INFO_COL_NAME = 4  # E
_STORE_INFO_HEADER_TOKENS = frozenset({"code", "store code", "kode"})


def get_store_name_map() -> dict:
    """code (col C) → full name (col E) from Store_Info. Cached 5 min."""
    cached = _cache_get("store_name_map")
    if cached is not None:
        return cached
    store_name_map: dict[str, str] = {}
    try:
        from services.auth_logic import get_sheet_client
        sh = get_sheet_client()
        ws = sh.worksheet("Store_Info")
        rows = ws.get_all_values()
        for row in rows[1:]:
            if len(row) <= _STORE_INFO_COL_CODE:
                continue
            code = str(row[_STORE_INFO_COL_CODE]).strip()
            if not code or code.lower() in _STORE_INFO_HEADER_TOKENS:
                continue
            name = (
                str(row[_STORE_INFO_COL_NAME]).strip()
                if len(row) > _STORE_INFO_COL_NAME
                else ""
            )
            store_name_map[code] = name or code
    except Exception:
        store_name_map = {}
    return _cache_set("store_name_map", store_name_map)


def get_at1_store_codes() -> list[str]:
    """Return store codes from Store_Info column C (same sheet as name map). Cached 5 min."""
    cached = _cache_get("at1_store_codes")
    if cached is not None:
        return cached
    try:
        from services.auth_logic import get_sheet_client
        sh = get_sheet_client()
        ws = sh.worksheet("Store_Info")
        rows = ws.get_all_values()
        if not rows:
            return _cache_set("at1_store_codes", [])
        seen: set[str] = set()
        for row in rows[1:]:
            if len(row) <= _STORE_INFO_COL_CODE:
                continue
            code = str(row[_STORE_INFO_COL_CODE]).strip()
            if not code or code.lower() in _STORE_INFO_HEADER_TOKENS:
                continue
            seen.add(code)
        stores = sorted(seen)
        return _cache_set("at1_store_codes", stores)
    except Exception:
        return _cache_set("at1_store_codes", [])


def get_sku_photo_map(db: Session, skus: set) -> dict:
    """
    Return sku → product_picture URL by reverse-lookup through PidStoreMap.
    PidStoreMap.sku may be a bundle (e.g. 'SKUA+SKUB'), so we parse tokens.
    """
    if not skus:
        return {}
    from sqlalchemy import or_, tuple_ as sql_tuple

    skus = {str(s).strip() for s in skus if s}
    token_to_key: dict[str, tuple] = {}

    # Fast path: exact match on pid_store_map.sku (single-SKU rows).
    sku_list = list(skus)
    chunk_size = 400
    for i in range(0, len(sku_list), chunk_size):
        chunk = sku_list[i : i + chunk_size]
        exact_rows = db.query(PidStoreMap).filter(PidStoreMap.sku.in_(chunk)).all()
        for pm in exact_rows:
            for token in parse_sku_tokens(pm.sku or ""):
                if token in skus and token not in token_to_key:
                    token_to_key[token] = (pm.store, pm.pid)

    remaining = {s for s in skus if s not in token_to_key}
    if remaining:
        # Bundle / combined SKU strings only (avoid full-table scan).
        bundle_rows = db.query(PidStoreMap).filter(
            PidStoreMap.sku.isnot(None),
            or_(
                PidStoreMap.sku.contains("+"),
                PidStoreMap.sku.contains(","),
                PidStoreMap.sku.contains("|"),
            ),
        ).all()
        for pm in bundle_rows:
            for token in parse_sku_tokens(pm.sku or ""):
                if token in remaining and token not in token_to_key:
                    token_to_key[token] = (pm.store, pm.pid)

    if not token_to_key:
        return {}

    # Batch-query ProductPerformance for product pictures
    pairs = list(token_to_key.values())
    pic_rows = db.query(
        ProductPerformance.store,
        ProductPerformance.pid,
        ProductPerformance.product_picture,
    ).filter(
        sql_tuple(ProductPerformance.store, ProductPerformance.pid).in_(pairs),
        ProductPerformance.product_picture.isnot(None),
        ProductPerformance.product_picture != "",
    ).all()

    pid_to_pic: dict[tuple, str] = {(r.store, r.pid): r.product_picture for r in pic_rows}
    return {
        token: pid_to_pic[key]
        for token, key in token_to_key.items()
        if key in pid_to_pic
    }


def get_pid_photo_map(db: Session, pids: set, store: str | None = None) -> dict:
    """
    Return pid → product_picture URL from ProductPerformance table.
    Optionally filtered by store. Returns the first non-null picture per pid.
    """
    if not pids:
        return {}
    q = db.query(
        ProductPerformance.pid,
        ProductPerformance.product_picture,
    ).filter(
        ProductPerformance.pid.in_(pids),
        ProductPerformance.product_picture.isnot(None),
        ProductPerformance.product_picture != "",
    )
    if store:
        q = q.filter(ProductPerformance.store == store)
    result: dict[str, str] = {}
    for row in q.all():
        if row.pid not in result:
            result[row.pid] = row.product_picture
    return result


# ─────────────────────────────────────────────────────────────────
# SHOPEE PROCESSING
# ─────────────────────────────────────────────────────────────────

def process_shopee_upload(db: Session, file_bytes: bytes, week_num: int) -> dict:
    """
    Parse Shopee product performance Excel, map store via PID, save to DB.
    Returns preview rows + summary.
    """
    week_start, week_end = get_week_dates(week_num)
    week_label = f"Week {week_num}"

    store_map = get_store_mapping(db)

    df = pd.read_excel(io.BytesIO(file_bytes), header=0)
    header_text = " ".join(str(c).lower() for c in df.columns if c is not None)
    if "store" in header_text or "店" in header_text:
        raise ValueError("This file looks like TikTok data. Please choose TikTok platform.")

    # Shopee column positions (0-indexed)
    COL_PID        = 0
    COL_PICTURE    = 1
    COL_NAME       = 2
    COL_UNIT       = 3
    COL_GMV        = 4
    COL_IMPRESSION = 6
    COL_VISITOR    = 7
    COL_CLICK      = 8
    COL_CTR        = 9
    COL_CO         = 14

    records = []
    skipped = 0
    for _, row in df.iterrows():
        pid = str(row.iloc[COL_PID]).strip() if pd.notna(row.iloc[COL_PID]) else ""
        if not pid or pid == "nan":
            skipped += 1
            continue

        store = store_map.get(pid, "-")

        visitor = float(row.iloc[COL_VISITOR]) if pd.notna(row.iloc[COL_VISITOR]) else 0
        unit = float(row.iloc[COL_UNIT]) if pd.notna(row.iloc[COL_UNIT]) else 0
        co = (unit / visitor) if visitor else 0

        records.append(ProductPerformance(
            week=week_label,
            week_start=week_start,
            week_end=week_end,
            platform="Shopee",
            store=store,
            pid=pid,
            product_picture=str(row.iloc[COL_PICTURE]) if pd.notna(row.iloc[COL_PICTURE]) else "",
            product_name="",
            impression=float(row.iloc[COL_IMPRESSION]) if pd.notna(row.iloc[COL_IMPRESSION]) else 0,
            visitor=visitor,
            click=float(row.iloc[COL_CLICK]) if pd.notna(row.iloc[COL_CLICK]) else 0,
            unit=unit,
            gmv=float(row.iloc[COL_GMV]) if pd.notna(row.iloc[COL_GMV]) else 0,
            ctr=float(row.iloc[COL_CTR]) if pd.notna(row.iloc[COL_CTR]) else 0,
            co=co,
        ))

    db.bulk_save_objects(records)
    db.commit()

    preview = [
        {
            "pid": r.pid,
            "store": r.store,
            "product_name": r.product_name,
            "product_picture": r.product_picture,
            "unit": r.unit,
            "gmv": r.gmv,
            "impression": r.impression,
            "visitor": r.visitor,
            "click": r.click,
            "ctr": r.ctr,
            "co": r.co,
        }
        for r in records[:50]
    ]

    return {
        "saved": len(records),
        "skipped": skipped,
        "week": week_label,
        "week_start": week_start,
        "week_end": week_end,
        "preview": preview,
    }


def _excel_identifier_str(val: object) -> str:
    """Normalize store / PID / MID from Excel without float/scientific corruption for large integers."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, str):
        s = val.strip()
        if s.lower() in ("nan", "none", ""):
            return ""
        if s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]
        return s
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val.is_integer() and abs(val) < 2**53:
            return str(int(val))
        return str(val).strip()
    return str(val).strip()


def _metric_key(header: object) -> str | None:
    text = str(header).strip().lower() if header is not None else ""
    if "gmv" in text or "销售额" in text:
        return "gmv"
    if "item sold" in text or "销量" in text:
        return "unit"
    if "impression" in text or "曝光" in text:
        return "impression"
    if text == "pv" or text.endswith(" pv") or text.startswith("pv ") or " pv " in text:
        return "pv"
    if text == "uv" or text.endswith(" uv") or text.startswith("uv ") or " uv " in text:
        return "uv"
    return None


def process_tiktok_upload(db: Session, file_bytes: bytes, week_num: int) -> dict:
    """
    Parse TikTok performance Excel (TT Data New). Store in product_performance.
    Columns: A=Store, B=PID, C=Picture. Skip D/E/F summary.
    Metrics appear in repeated groups; sum by header name.
    """
    pid_to_store = get_pid_to_store_map(db)
    pid_to_sku = get_pid_to_sku_map(db)
    store_pid_to_sku = get_store_pid_to_sku_map(db)
    week_start, week_end = get_week_dates(week_num)
    week_label = f"Week {week_num}"

    raw_df = pd.read_excel(io.BytesIO(file_bytes), header=None)
    if raw_df.empty:
        return {
            "saved": 0,
            "skipped": 0,
            "week": week_label,
            "week_start": week_start,
            "week_end": week_end,
            "preview": [],
        }

    header_row = 0
    for i in range(min(10, len(raw_df))):
        row_vals = [str(v).strip().lower() for v in raw_df.iloc[i].tolist() if v is not None]
        if any("store" in v or "店" in v for v in row_vals) and any(
            "pid" in v or "product" in v or "产品id" in v or "商品id" in v for v in row_vals
        ):
            header_row = i
            break

    def _norm(val: object) -> str:
        return "".join(ch for ch in str(val).lower() if ch.isalnum()) if val is not None else ""

    header_only = pd.read_excel(io.BytesIO(file_bytes), header=header_row, nrows=0)
    store_col = None
    pid_col = None
    picture_col = None
    for col in header_only.columns:
        n = _norm(col)
        col_l = str(col).lower()
        if store_col is None and ("store" in n or "店" in str(col)):
            store_col = col
        if pid_col is None and (
            "pid" in n or "productid" in n or "商品id" in col_l or "产品id" in col_l
        ):
            pid_col = col
        if picture_col is None and ("picture" in n or "image" in n or "图片" in str(col)):
            picture_col = col

    str_cols = {c for c in (store_col, pid_col, picture_col) if c is not None}
    dtype_arg = {c: str for c in str_cols} if str_cols else None
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_row, dtype=dtype_arg)
    metric_cols = {"gmv": [], "unit": [], "impression": [], "pv": [], "uv": []}
    for idx, col in enumerate(df.columns):
        if idx in (3, 4, 5):
            continue
        key = _metric_key(col)
        if key:
            metric_cols[key].append(col)

    records = []
    skipped = 0
    for _, row in df.iterrows():
        store_val = row[store_col] if store_col in row else (row.iloc[0] if len(row) > 0 else None)
        pid_val = row[pid_col] if pid_col in row else (row.iloc[1] if len(row) > 1 else None)
        pic_val = row[picture_col] if picture_col in row else (row.iloc[2] if len(row) > 2 else None)

        store = _excel_identifier_str(store_val)
        pid = _excel_identifier_str(pid_val)
        picture = _excel_identifier_str(pic_val)
        if not store or not pid or store == "nan" or pid == "nan":
            skipped += 1
            continue

        gmv = sum(float(row[c]) for c in metric_cols["gmv"] if pd.notna(row[c])) if metric_cols["gmv"] else 0
        unit = sum(float(row[c]) for c in metric_cols["unit"] if pd.notna(row[c])) if metric_cols["unit"] else 0
        impression = sum(float(row[c]) for c in metric_cols["impression"] if pd.notna(row[c])) if metric_cols["impression"] else 0
        pv = sum(float(row[c]) for c in metric_cols["pv"] if pd.notna(row[c])) if metric_cols["pv"] else 0
        uv = sum(float(row[c]) for c in metric_cols["uv"] if pd.notna(row[c])) if metric_cols["uv"] else 0

        ctr = (pv / impression) if impression else 0
        co = (unit / uv) if uv else 0

        if store in ("", "-", "—"):
            store = pid_to_store.get(pid, "-")
        product_name = store_pid_to_sku.get((store, pid)) or pid_to_sku.get(pid, "")

        records.append(ProductPerformance(
            week=week_label,
            week_start=week_start,
            week_end=week_end,
            platform="TikTok",
            store=store,
            pid=pid,
            product_picture=picture,
            product_name=product_name,
            impression=impression,
            visitor=uv,
            click=pv,
            unit=unit,
            gmv=gmv,
            ctr=ctr,
            co=co,
        ))

    db.bulk_save_objects(records)
    db.commit()

    preview = [
        {
            "pid": r.pid,
            "store": r.store,
            "unit": r.unit,
            "gmv": r.gmv,
            "impression": r.impression,
            "visitor": r.visitor,
            "click": r.click,
            "ctr": r.ctr,
            "co": r.co,
            "product_picture": r.product_picture,
            "product_name": r.product_name,
        }
        for r in records[:50]
    ]

    return {
        "saved": len(records),
        "skipped": skipped,
        "week": week_label,
        "week_start": week_start,
        "week_end": week_end,
        "preview": preview,
    }


# ─────────────────────────────────────────────────────────────────
# CONVERTER UPDATE — Shopee
# ─────────────────────────────────────────────────────────────────

def _is_tiktok_store(store: str) -> bool:
    return store.strip().upper().startswith("TT")


def _parse_converter_shopee(file_bytes: bytes) -> list[dict]:
    # Load workbook (openpyxl Set descriptor already patched at module level)
    wb = _load_workbook_lenient(file_bytes)
    ws = wb.active
    # Skip first 6 rows (header metadata), data starts row 7
    all_rows = list(ws.iter_rows(min_row=7, values_only=True))
    wb.close()

    # File columns: A=PID, B=Product Name, C=Variation ID (MID), D=Variation Name, E=Parent SKU, F=SKU, G=SKU fallback
    COL_PID = 0
    COL_MID = 2
    COL_SKU = 5
    COL_SKU_FALLBACK = 6

    parsed = []
    for row in all_rows:
        pid = str(row[COL_PID]).strip() if row[COL_PID] is not None else ""
        if not pid or pid == "None":
            continue

        mid = str(row[COL_MID]).strip() if len(row) > COL_MID and row[COL_MID] is not None else ""

        sku_raw = row[COL_SKU] if len(row) > COL_SKU else None
        sku = str(sku_raw).strip() if sku_raw is not None and str(sku_raw).strip() not in ("", "None") else ""
        if not sku:
            sku_fb = row[COL_SKU_FALLBACK] if len(row) > COL_SKU_FALLBACK else None
            sku = str(sku_fb).strip() if sku_fb is not None and str(sku_fb).strip() not in ("", "None") else ""

        parsed.append({"pid": pid, "mid": mid, "sku": sku})

    return [item for item in parsed if item["mid"]]


def _parse_converter_tiktok(file_bytes: bytes) -> list[dict]:
    try:
        wb = _load_workbook_lenient(file_bytes)
        ws = wb.active
        header_row, col_map = _find_tiktok_header_openpyxl(ws)
        data_start = header_row + 1 if header_row else 6
        all_rows = list(ws.iter_rows(min_row=data_start, values_only=True))
        wb.close()

        # Fallbacks if headers are missing
        col_pid = (col_map.get("pid") or 1) - 1
        col_mid = (col_map.get("mid") or 4) - 1
        col_sku = (col_map.get("seller_sku") or 12) - 1

        pid_headers = {"productid", "pid"}
        mid_headers = {"skuid", "skuitemid", "variationid", "mid"}
        sku_headers = {"sellersku"}

        parsed = []
        for row in all_rows:
            pid = str(row[col_pid]).strip() if len(row) > col_pid and row[col_pid] is not None else ""
            if not pid or pid == "None":
                continue

            mid = str(row[col_mid]).strip() if len(row) > col_mid and row[col_mid] is not None else ""
            if not mid or mid == "None":
                continue

            sku_raw = row[col_sku] if len(row) > col_sku else None
            sku = str(sku_raw).strip() if sku_raw is not None and str(sku_raw).strip() not in ("", "None") else ""

            if _normalize_header(pid) in pid_headers or _normalize_header(mid) in mid_headers or _normalize_header(sku) in sku_headers:
                continue

            parsed.append({"pid": pid, "mid": mid, "sku": sku})

        return parsed
    except Exception:
        return _parse_xlsx_minimal_tiktok(file_bytes)


def _parse_xlsx_minimal_tiktok(file_bytes: bytes) -> list[dict]:
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        sheet_name = "xl/worksheets/sheet1.xml"
        if sheet_name not in zf.namelist():
            sheets = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
            if not sheets:
                return []
            sheet_name = sorted(sheets)[0]

        shared = _read_shared_strings(zf)
        sheet_xml = zf.read(sheet_name)

    header_rows = _read_sheet_rows(sheet_xml, shared, max_row=10)
    header_row, col_map = _find_tiktok_header_rows(header_rows)
    data_start = (header_row + 1) if header_row else 6

    if "pid" not in col_map:
        col_map["pid"] = 1
    if "mid" not in col_map:
        col_map["mid"] = 4
    if "seller_sku" not in col_map:
        col_map["seller_sku"] = 12

    return _read_sheet_columns(sheet_xml, shared, data_start, {
        "pid": col_map["pid"],
        "mid": col_map["mid"],
        "sku": col_map["seller_sku"],
    })


def _read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        xml_bytes = zf.read("xl/sharedStrings.xml")
    except Exception:
        return []

    root = ET.fromstring(xml_bytes)
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    shared = []
    for si in root.findall(f"{ns}si"):
        texts = [t.text or "" for t in si.findall(f".//{ns}t")]
        shared.append("".join(texts))
    return shared


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value).lower() if ch.isalnum())


def _find_tiktok_header_openpyxl(ws) -> tuple[int | None, dict]:
    pid_headers = {"productid", "pid"}
    mid_headers = {"skuid", "skuitemid", "variationid", "mid"}
    sku_headers = {"sellersku"}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        col_map = {}
        for col_idx, cell in enumerate(row, start=1):
            norm = _normalize_header(cell)
            if norm in pid_headers:
                col_map["pid"] = col_idx
            if norm in mid_headers:
                col_map["mid"] = col_idx
            if norm in sku_headers:
                col_map["seller_sku"] = col_idx
        if "seller_sku" in col_map:
            return row_idx, col_map
    return None, {}


def _find_tiktok_header_rows(rows: list[tuple[int, dict]]) -> tuple[int | None, dict]:
    pid_headers = {"productid", "pid"}
    mid_headers = {"skuid", "skuitemid", "variationid", "mid"}
    sku_headers = {"sellersku"}
    for row_idx, cells in rows:
        col_map = {}
        for col_idx, cell in cells.items():
            norm = _normalize_header(cell)
            if norm in pid_headers:
                col_map["pid"] = col_idx
            if norm in mid_headers:
                col_map["mid"] = col_idx
            if norm in sku_headers:
                col_map["seller_sku"] = col_idx
        if "seller_sku" in col_map:
            return row_idx, col_map
    return None, {}


def _cell_ref_to_col_row(cell_ref: str) -> tuple[int, int]:
    col = 0
    row = 0
    for ch in cell_ref:
        if "A" <= ch <= "Z":
            col = col * 26 + (ord(ch) - ord("A") + 1)
        elif "0" <= ch <= "9":
            row = row * 10 + (ord(ch) - ord("0"))
    return col, row


def _read_sheet_rows(sheet_xml: bytes, shared: list[str], min_row: int = 1, max_row: int | None = None) -> list[tuple[int, dict]]:
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    root = ET.fromstring(sheet_xml)
    rows = []
    for row in root.findall(f".//{ns}row"):
        row_idx = int(row.attrib.get("r", "0") or "0")
        if row_idx < min_row:
            continue
        if max_row is not None and row_idx > max_row:
            continue

        values = {}
        for cell in row.findall(f"{ns}c"):
            ref = cell.attrib.get("r", "")
            col_idx, _ = _cell_ref_to_col_row(ref)
            cell_type = cell.attrib.get("t")
            v = cell.find(f"{ns}v")
            if cell_type == "s" and v is not None:
                idx = int(v.text) if v.text is not None else -1
                cell_val = shared[idx] if 0 <= idx < len(shared) else ""
            elif cell_type == "inlineStr":
                t = cell.find(f".//{ns}t")
                cell_val = t.text if t is not None else ""
            else:
                cell_val = v.text if v is not None else ""

            values[col_idx] = (cell_val or "").strip()

        rows.append((row_idx, values))

    return rows


def _read_sheet_columns(sheet_xml: bytes, shared: list[str], min_row: int, col_map: dict) -> list[dict]:
    pid_headers = {"productid", "pid"}
    mid_headers = {"skuid", "skuitemid", "variationid", "mid"}
    sku_headers = {"sellersku"}
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    root = ET.fromstring(sheet_xml)
    rows = []
    for row in root.findall(f".//{ns}row"):
        row_idx = int(row.attrib.get("r", "0") or "0")
        if row_idx < min_row:
            continue

        values = {"pid": "", "mid": "", "sku": ""}
        for cell in row.findall(f"{ns}c"):
            ref = cell.attrib.get("r", "")
            col_idx, _ = _cell_ref_to_col_row(ref)
            if col_idx not in col_map.values():
                continue

            cell_type = cell.attrib.get("t")
            v = cell.find(f"{ns}v")
            if cell_type == "s" and v is not None:
                idx = int(v.text) if v.text is not None else -1
                cell_val = shared[idx] if 0 <= idx < len(shared) else ""
            elif cell_type == "inlineStr":
                t = cell.find(f".//{ns}t")
                cell_val = t.text if t is not None else ""
            else:
                cell_val = v.text if v is not None else ""

            for key, target_col in col_map.items():
                if col_idx == target_col:
                    values[key] = (cell_val or "").strip()

        if values["pid"] and values["mid"]:
            if _normalize_header(values["pid"]) in pid_headers or _normalize_header(values["mid"]) in mid_headers or _normalize_header(values["sku"]) in sku_headers:
                continue
            rows.append(values)

    return rows


def _load_workbook_lenient(file_bytes: bytes):
    try:
        return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)


def _extract_excel_files(file_bytes: bytes, filename: str | None) -> list[tuple[str, bytes]]:
    fname_lower = (filename or "").lower()
    is_zip_ext = fname_lower.endswith(".zip")

    if is_zip_ext:
        if not zipfile.is_zipfile(io.BytesIO(file_bytes)):
            return []
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            names = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xls")) and not n.endswith("/")]
            return [(name, zf.read(name)) for name in sorted(names)]

    # .xlsx/.xls — treat as direct Excel file (xlsx is internally a zip, don't unpack it)
    return [(filename or "upload.xlsx", file_bytes)]


def process_converter_upload(db: Session, file_bytes: bytes, store: str, filename: str | None = None) -> dict:
    """
    Parse converter Excel and upsert pid_store_map by (store, mid).

    Shopee format:
      Row 7+, Col A=PID, Col C=MID, Col F=SKU (fallback Col G)
    TikTok format:
      Row 6+, Col A=PID, Col D=MID, Col L=SKU
    """
    from datetime import datetime

    files = _extract_excel_files(file_bytes, filename)
    if not files:
        raise ValueError("No Excel files found in zip")

    parsed = []
    errors = []
    for name, data in files:
        try:
            if _is_tiktok_store(store):
                parsed.extend(_parse_converter_tiktok(data))
            else:
                parsed.extend(_parse_converter_shopee(data))
        except Exception as exc:
            errors.append(f"{name}: {exc}")

    if not parsed and errors:
        raise ValueError("; ".join(errors[:3]))

    deduped = {}
    for item in parsed:
        mid = item.get("mid") or ""
        if not mid:
            continue
        deduped[mid] = item

    parsed = list(deduped.values())

    if not parsed:
        return {"saved": 0, "deleted": 0, "unique_pids": 0, "unique_mids": 0, "store": store}

    now = datetime.utcnow()
    for item in parsed:
        existing = db.query(PidStoreMap).filter(
            PidStoreMap.store == store,
            PidStoreMap.mid == item["mid"],
        ).first()
        if existing:
            existing.pid = item["pid"]
            existing.sku = item["sku"]
            existing.updated_at = now
        else:
            db.add(PidStoreMap(
                mid=item["mid"],
                pid=item["pid"],
                store=store,
                sku=item["sku"],
                updated_at=now,
            ))
    db.commit()

    unique_pids = len({r["pid"] for r in parsed})
    unique_mids = len({r["mid"] for r in parsed})

    return {
        "saved": len(parsed),
        "deleted": 0,
        "unique_pids": unique_pids,
        "unique_mids": unique_mids,
        "store": store,
        "preview": parsed[:30],
    }


def get_converter_stats(db: Session) -> dict:
    """Per-store stats: unique PIDs, total MIDs (rows), last updated."""
    from sqlalchemy import func as sqlfunc
    store_name_map = get_store_name_map()

    rows = (
        db.query(
            PidStoreMap.store,
            sqlfunc.count(sqlfunc.distinct(PidStoreMap.pid)).label("pid_count"),
            sqlfunc.count(PidStoreMap.mid).label("mid_count"),
            sqlfunc.max(PidStoreMap.updated_at).label("last_updated"),
        )
        .group_by(PidStoreMap.store)
        .order_by(PidStoreMap.store)
        .all()
    )
    stats = []
    shopee_count = 0
    tiktok_count = 0
    for r in rows:
        store_code = r.store
        if store_code and store_code.upper().startswith("TT"):
            tiktok_count += 1
        else:
            shopee_count += 1

        stats.append({
            "store_code": store_code,
            "store_name": store_name_map.get(store_code, store_code),
            "pid_count": r.pid_count,
            "mid_count": r.mid_count,
            "last_updated": r.last_updated.isoformat() if r.last_updated else None,
        })

    return {
        "summary": {
            "shopee": shopee_count,
            "tiktok": tiktok_count,
        },
        "rows": stats,
    }


# ─────────────────────────────────────────────────────────────────
# SKU BRAND AGGREGATION
# ─────────────────────────────────────────────────────────────────

import re as _re

_SKU_DELIMITERS = r"[+/\-|,;\" ]"
_MIN_SKU_LEN = 12


def parse_sku_tokens(sku_string: str) -> list[str]:
    """
    Split a SKU/bundle string by delimiters and return only tokens
    that are valid SKUs (length >= 12).
    """
    if not sku_string or not sku_string.strip():
        return []
    parts = _re.split(_SKU_DELIMITERS, sku_string.strip())
    return [p.strip() for p in parts if len(p.strip()) >= _MIN_SKU_LEN]


def compute_sku_performance(db: Session, week: str, platform: str) -> dict:
    """
    Aggregate product_performance rows into sku_performance.
    For each PID row:
      - Look up SKU from converter (store_pid_to_sku or pid_to_sku)
      - Parse SKU tokens (handle bundles)
      - Each token gets full metrics of the parent row
    Aggregate per (week, platform, store, sku):
      - Sum: gmv, unit, impression, visitor, click
      - CTR = sum(click) / sum(impression)
      - CO  = sum(unit)  / sum(visitor)
      - pid_count = distinct PIDs that contributed to this SKU
    Returns summary dict.
    """
    from models import SkuPerformance, FreemirPrice

    store_pid_to_sku = get_store_pid_to_sku_map(db)
    pid_to_sku = get_pid_to_sku_map(db)

    # Only keep SKUs that exist in the Price master (freemir_price)
    valid_skus: set[str] = {r.sku for r in db.query(FreemirPrice.sku).all()}

    # Load all performance rows for this week + platform
    rows = (
        db.query(ProductPerformance)
        .filter(ProductPerformance.week == week, ProductPerformance.platform == platform)
        .all()
    )

    if not rows:
        return {"computed": 0, "week": week, "platform": platform}

    # Use the first row to get week_start / week_end
    week_start = rows[0].week_start
    week_end = rows[0].week_end

    # Aggregate: key = (store, sku)
    agg: dict[tuple[str, str], dict] = {}

    for r in rows:
        sku_raw = store_pid_to_sku.get((r.store, r.pid)) or pid_to_sku.get(r.pid, "")
        tokens = parse_sku_tokens(sku_raw)
        if not tokens:
            continue  # no valid SKU mapping for this PID — skip (don't fall back to PID)

        for token in tokens:
            # Skip tokens that are not in the Price master whitelist
            if valid_skus and token not in valid_skus:
                continue
            key = (r.store, token)
            if key not in agg:
                agg[key] = {
                    "store": r.store,
                    "sku": token,
                    "impression": 0.0,
                    "visitor": 0.0,
                    "click": 0.0,
                    "unit": 0.0,
                    "gmv": 0.0,
                    "pids": set(),
                }
            bucket = agg[key]
            bucket["impression"] += r.impression or 0
            bucket["visitor"]   += r.visitor    or 0
            bucket["click"]     += r.click      or 0
            bucket["unit"]      += r.unit       or 0
            bucket["gmv"]       += r.gmv        or 0
            bucket["pids"].add(r.pid)

    # Delete existing sku_performance rows for this week + platform
    db.query(SkuPerformance).filter(
        SkuPerformance.week == week,
        SkuPerformance.platform == platform,
    ).delete()

    # Insert aggregated rows
    objects = []
    for (store, sku), bucket in agg.items():
        imp = bucket["impression"]
        vis = bucket["visitor"]
        clk = bucket["click"]
        unt = bucket["unit"]
        objects.append(SkuPerformance(
            week=week,
            week_start=week_start,
            week_end=week_end,
            platform=platform,
            store=store,
            sku=sku,
            impression=imp,
            visitor=vis,
            click=clk,
            unit=unt,
            gmv=bucket["gmv"],
            ctr=(clk / imp) if imp else 0.0,
            co=(unt / vis) if vis else 0.0,
            pid_count=len(bucket["pids"]),
        ))

    db.bulk_save_objects(objects)
    db.commit()

    return {"computed": len(objects), "week": week, "platform": platform}


# ─────────────────────────────────────────────────────────────────
# SKU COMPARISON
# ─────────────────────────────────────────────────────────────────

def get_sku_comparison(db: Session, week_a: str, week_b: str, platform: str = "All") -> dict:
    """
    Compare SKU performance between two weeks.
    Returns sections: All Brand, per Platform (if All), per Store.
    """
    from models import SkuPerformance

    store_name_map = get_store_name_map()

    def _load(week: str):
        q = db.query(SkuPerformance).filter(SkuPerformance.week == week)
        if platform and platform.lower() != "all":
            q = q.filter(SkuPerformance.platform == platform)
        return q.all()

    rows_a = _load(week_a)
    rows_b = _load(week_b)

    # Extract date ranges from the loaded rows
    def _dates(rows, week):
        for r in rows:
            if r.week_start and r.week_end:
                return r.week_start, r.week_end
        # Fallback: compute from week number
        import re as _re2
        m = _re2.search(r'(\d+)', week)
        if m:
            ws, we = get_week_dates(int(m.group(1)))
            return ws, we
        return None, None

    date_a = _dates(rows_a, week_a)
    date_b = _dates(rows_b, week_b)

    def _empty():
        return {"impression": 0.0, "visitor": 0.0, "click": 0.0, "unit": 0.0, "gmv": 0.0}

    def _ctr_co(m):
        m["ctr"] = (m["click"] / m["impression"]) if m["impression"] else 0.0
        m["co"]  = (m["unit"]  / m["visitor"])    if m["visitor"]    else 0.0
        return m

    def _pct(a, b):
        if not a:
            return None
        return round((b - a) / a * 100, 2)

    def _growth_dict(ta, tb):
        return {
            "impression": _pct(ta["impression"], tb["impression"]),
            "visitor":    _pct(ta["visitor"],    tb["visitor"]),
            "click":      _pct(ta["click"],      tb["click"]),
            "unit":       _pct(ta["unit"],       tb["unit"]),
            "gmv":        _pct(ta["gmv"],        tb["gmv"]),
            "gmv_gap":    round(tb["gmv"] - ta["gmv"], 0),
            "ctr":        _pct(ta.get("ctr", 0), tb.get("ctr", 0)),
            "co":         _pct(ta.get("co", 0),  tb.get("co", 0)),
        }

    def _agg_by_sku(rows):
        d: dict = {}
        pid_d: dict = {}  # sku -> set of pid_count (approximate sum)
        for r in rows:
            sku = r.sku or ""
            if sku not in d:
                d[sku] = _empty()
                pid_d[sku] = 0
            for k in ("impression", "visitor", "click", "unit", "gmv"):
                d[sku][k] += getattr(r, k) or 0.0
            pid_d[sku] += r.pid_count or 0
        for m in d.values():
            _ctr_co(m)
        return d, pid_d

    def _sum_metrics(rows):
        t = _empty()
        for r in rows:
            for k in ("impression", "visitor", "click", "unit", "gmv"):
                t[k] += getattr(r, k) or 0.0
        return _ctr_co(t)

    def _top_store_by_metric(rows, metric: str):
        per_store = {}
        for r in rows:
            sku = r.sku or ""
            if not sku:
                continue
            store = (r.store or "").strip()
            if not store or store in ("-", "—"):
                continue
            platform = (r.platform or "").strip()
            key = (platform, store)
            if sku not in per_store:
                per_store[sku] = {}
            per_store[sku][key] = per_store[sku].get(key, 0.0) + (getattr(r, metric) or 0.0)
        top_map = {}
        for sku, store_vals in per_store.items():
            top_key = max(store_vals.items(), key=lambda x: x[1])[0] if store_vals else None
            if not top_key:
                top_map[sku] = None
                continue
            plat, store = top_key
            store_label = store_name_map.get(store, store)
            top_map[sku] = f"{plat} {store_label}".strip() if plat else store_label
        return top_map

    def _make_section(section_id: str, store_code, rows_a_sub, rows_b_sub):
        map_a, pid_a = _agg_by_sku(rows_a_sub)
        map_b, pid_b = _agg_by_sku(rows_b_sub)
        # Highest-store labels use Period B as primary baseline, fallback to Period A.
        gmv_top_b = _top_store_by_metric(rows_b_sub, "gmv")
        gmv_top_a = _top_store_by_metric(rows_a_sub, "gmv")
        imp_top_b = _top_store_by_metric(rows_b_sub, "impression")
        imp_top_a = _top_store_by_metric(rows_a_sub, "impression")
        all_skus = sorted(
            set(map_a) | set(map_b),
            key=lambda s: -(map_a.get(s, {}).get("gmv", 0) + map_b.get(s, {}).get("gmv", 0)),
        )
        total_a = _sum_metrics(rows_a_sub)
        total_b = _sum_metrics(rows_b_sub)
        total_pid_a = sum(pid_a.values())
        total_pid_b = sum(pid_b.values())

        sku_rows = []
        for sku in all_skus:
            a = {**_empty(), "ctr": 0.0, "co": 0.0}
            a.update(map_a.get(sku, {}))
            b = {**_empty(), "ctr": 0.0, "co": 0.0}
            b.update(map_b.get(sku, {}))
            sku_rows.append({
                "sku": sku,
                "a": a, "b": b,
                "growth": _growth_dict(a, b),
                "pid_a": pid_a.get(sku, 0),
                "pid_b": pid_b.get(sku, 0),
                "highest_gmv_store": gmv_top_b.get(sku) or gmv_top_a.get(sku),
                "highest_impression_store": imp_top_b.get(sku) or imp_top_a.get(sku),
            })

        return {
            "section": section_id,
            "store_code": store_code,
            "store_name": store_name_map.get(store_code, section_id) if store_code else section_id,
            "total_a": total_a,
            "total_b": total_b,
            "total_pid_a": total_pid_a,
            "total_pid_b": total_pid_b,
            "growth": _growth_dict(total_a, total_b),
            "rows": sku_rows,
        }

    sections = []

    # 1. All Brand
    sections.append(_make_section("All Brand", None, rows_a, rows_b))

    # 2. Per Platform (only when platform = All)
    if platform.lower() == "all":
        all_platforms = sorted(set(r.platform for r in rows_a + rows_b if r.platform))
        for plat in all_platforms:
            sections.append(_make_section(
                plat, None,
                [r for r in rows_a if r.platform == plat],
                [r for r in rows_b if r.platform == plat],
            ))

    # 3. Per Store
    all_stores = sorted(set(
        r.store for r in rows_a + rows_b if r.store and r.store not in ("", "-", "\u2014")
    ))
    for store in all_stores:
        sections.append(_make_section(
            store, store,
            [r for r in rows_a if r.store == store],
            [r for r in rows_b if r.store == store],
        ))

    # Collect all SKUs, then build a name/photo map from freemir_name + product_performance
    from models import FreemirName
    all_sku_set: set[str] = set()
    for sec in sections:
        for row in sec["rows"]:
            if row["sku"]:
                all_sku_set.add(row["sku"])

    name_map = {
        n.sku: {"product_name": n.product_name, "product_link": n.link, "mark": n.mark}
        for n in db.query(FreemirName).filter(FreemirName.sku.in_(all_sku_set)).all()
    }
    photo_map = get_sku_photo_map(db, all_sku_set)

    for sec in sections:
        for row in sec["rows"]:
            nm = name_map.get(row["sku"], {})
            row["product_name"] = nm.get("product_name")
            row["product_link"] = nm.get("product_link")
            row["mark"] = nm.get("mark")
            # Prefer SKU_Info link as canonical SKU image, then fallback to PID-based snapshot.
            row["photo"] = nm.get("product_link") or photo_map.get(row["sku"])

    return {
        "week_a": week_a, "week_b": week_b, "platform": platform,
        "week_a_start": date_a[0], "week_a_end": date_a[1],
        "week_b_start": date_b[0], "week_b_end": date_b[1],
        "sections": sections,
    }

