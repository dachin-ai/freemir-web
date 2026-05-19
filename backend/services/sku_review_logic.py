"""
SKU Review Analysis — split bundle SKUs and extract issue/part keywords
with negation-aware phrase matching (e.g. "tidak bagus" ≠ positive "bagus").

Vocabulary is enriched from real after-sales exports:
  - Macro tags: ITEM_MISSING, SLIGHT_SCRATCH_DENTS, WRONG_ITEM,
    FUNCTIONAL_DMG, DIFFERENT_DESCRIPTION, BROKEN_PRODUCTS,
    SPILLED_CONTENTS, INCOMPLATE_PACKAGE
  - Common phrases: "doesn't fit", "can't work", "doesn't match description",
    "no longer needed", "missing screw / bolt / handle lid"
  - New part vocab: bolt, screw, knob, cable, steamer, spatula, stove, glass_lid
"""
from __future__ import annotations

import io
import re
from collections import Counter, defaultdict
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------
COLUMN_ALIASES = {
    "shop": ["shop"],
    "after_sales_order_no": ["after sales order no", "after-sales order no"],
    "order_serial_number": ["order serial number", "erp order", "serial number"],
    "order_id": ["order id", "platform order"],
    "business_type": ["business type"],
    "after_sales_type": ["after sales type", "after-sales type"],
    "sku": ["sku"],
    "detail": ["detail", "complaint", "remarks", "description"],
}


def _normalize_header(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip().lower())


def _map_columns(df: pd.DataFrame) -> Dict[str, str]:
    norm = {_normalize_header(c): c for c in df.columns}
    mapping: Dict[str, str] = {}
    for key, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in norm:
                mapping[key] = norm[alias]
                break
    missing = [k for k in ("sku", "detail") if k not in mapping]
    if missing:
        cols = ", ".join(map(str, df.columns.tolist()[:12]))
        raise ValueError(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Expected headers like Shop, SKU, Detail. Found: {cols}"
        )
    return mapping


def _read_upload(content: bytes, filename: str) -> pd.DataFrame:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        for enc in ("utf-8", "utf-8-sig", "latin-1", "gbk"):
            try:
                return pd.read_csv(io.BytesIO(content), dtype=str, encoding=enc)
            except Exception:
                continue
        raise ValueError("Could not read CSV file. Try saving as UTF-8.")
    if lower.endswith((".xlsx", ".xls")):
        xl = pd.ExcelFile(io.BytesIO(content))
        # Prefer dedicated result-style sheets if present
        sheet = xl.sheet_names[0]
        for cand in ("Result", "February", "March", "April", "May"):
            if cand in xl.sheet_names:
                sheet = cand
                break
        df = pd.read_excel(io.BytesIO(content), sheet_name=sheet, dtype=str)
        return df
    raise ValueError("Unsupported file type. Upload .xlsx or .csv.")


# ---------------------------------------------------------------------------
# Vocabulary
# ---------------------------------------------------------------------------
# Macro tags — common uppercase tokens used by after-sales staff.
# Each maps to (issues, parts). Skip negation-check; they are explicit tags.
MACRO_RULES: List[Tuple[str, List[str], List[str]]] = [
    ("item_missing",            ["missing"],                       []),
    ("incomplate_package",      ["missing"],                       ["box"]),
    ("incomplete_package",      ["missing"],                       ["box"]),
    ("missing_handle_lid",      ["missing"],                       ["handle", "lid"]),
    ("missing_lid_handle",      ["missing"],                       ["lid", "handle"]),
    ("missing_handle_pot",      ["missing"],                       ["handle", "pan_body"]),
    ("missing_lid_glass",       ["missing"],                       ["glass_lid"]),
    ("missing_lid",             ["missing"],                       ["lid"]),
    ("missing_handle",          ["missing"],                       ["handle"]),
    ("missing_screw",           ["missing"],                       ["screw"]),
    ("missing_bolt",            ["missing"],                       ["bolt"]),
    ("missing_knob",            ["missing"],                       ["knob"]),
    ("missing_spatula",         ["missing"],                       ["spatula"]),
    ("missing_steamer",         ["missing"],                       ["steamer"]),
    ("missing_cable",           ["missing"],                       ["cable"]),
    ("slight_scratch_dents",    ["scratch", "dent"],               []),
    ("scratch_dents",           ["scratch", "dent"],               []),
    ("wrong_item",              ["wrong_item"],                    []),
    ("wrong_sku",               ["wrong_item"],                    []),
    ("functional_dmg",          ["not_working"],                   []),
    ("functional_damage",       ["not_working"],                   []),
    ("product_cant_work",       ["not_working"],                   []),
    ("product_doesnt_work",     ["not_working"],                   []),
    ("different_description",   ["not_as_described"],              []),
    ("not_as_described",        ["not_as_described"],              []),
    ("doesnt_match",            ["not_as_described"],              []),
    ("broken_products",         ["damage"],                        []),
    ("broken_lid_glass",        ["damage"],                        ["glass_lid"]),
    ("broken_lid",              ["damage"],                        ["lid"]),
    ("broken_handle_lid",       ["damage"],                        ["handle", "lid"]),
    ("broken_handle",           ["damage"],                        ["handle"]),
    ("spilled_contents",        ["spillage"],                      []),
    ("buyer_no_longer_need",    ["change_of_mind"],                []),
    ("no_longer_needed",        ["change_of_mind"],                []),
]

# Issue keywords (negation-aware applied at scan time)
ISSUE_RULES: List[Tuple[str, List[str]]] = [
    ("not_fit", [
        "doesn't fit", "does not fit", "doesnt fit", "do not fit",
        "don't fit", "dont fit", "don’t fit",
        "can't fit", "cannot fit", "can not fit", "won't fit",
        "didn't fit", "did not fit", "doesn't close properly",
        "does not close properly", "won't close", "doesn't close",
        "don't fit properly", "does not fit properly", "doesnt fit properly",
        "lid doesn't fit", "lid not fit", "lid not close",
        "tutup tidak pas", "tutup tidak cocok", "tidak pas",
        "tidak cocok", "tidak masuk", "tidak menutup",
        "ga pas", "gak pas", "kurang pas", "doesn't grip",
        "screw doesn't fit", "screw don't fit",
        "handle is not the right size", "not the right size",
    ]),
    ("not_as_described", [
        "doesn't match description", "does not match description",
        "doesn't match the description", "not match description",
        "different from description", "different description",
        "not as described", "not as expected",
        "not according to expectations", "not according to description",
        "not according to order", "tidak sesuai pesanan",
        "item does not match", "does not match", "do not match",
        "size and color do not match", "size do not match",
        "tidak sesuai", "tidak sesuai deskripsi",
        "tidak sesuai gambar", "tidak sesuai ekspektasi",
        "berbeda dari deskripsi", "berbeda dari gambar",
        "different size", "different color", "different model",
        "ordered pink", "ordered white", "received white", "received pink",
        "smaller size", "quite small", "too small", "very small",
        "too short", "really short", "cable is short",
    ]),
    ("not_working", [
        "doesn't work", "does not work", "didn't work", "did not work",
        "not working", "can't work", "cannot work", "can not work",
        "stops working", "stopped working", "the heat doesn't work",
        "heat doesn't work", "tidak berfungsi", "tidak bisa dipakai",
        "tidak bisa digunakan", "tidak nyala", "rusak fungsi",
        "ga berfungsi", "gak berfungsi",
    ]),
    ("spillage", [
        "spilled", "spill", "spilling", "leaked out", "overflow",
        "tumpah", "meluap", "berceceran", "isi tumpah",
        "spilled_contents",
    ]),
    ("change_of_mind", [
        "no longer needed", "no longer need", "tidak diperlukan lagi",
        "tidak butuh lagi", "berubah pikiran", "salah beli",
        "ganti pikiran", "buyer remorse",
    ]),
    ("missing", [
        "item missing", "parts missing", "part missing", "not included",
        "wasn't included", "was not included", "not included in the shipment",
        "bonus wasn't included", "wasn't included in",
        "didn't receive", "did not receive", "didnt receive",
        "not received", "tidak menerima", "barang tidak lengkap",
        "kurang lengkap", "incomplete", "incomplate",
        "missing", "hilang", "belum ada", "tidak ada", "tidak diterima",
        "no bolts", "no bolt", "no screw", "no screws",
        "there's no lid", "there is no lid", "tidak ada tutup",
        "wasn't sent", "was not sent", "didnt sent", "tidak dikirim",
        "kurang barang", "sent without", "without a handle", "without handle",
        "no handle", "belum dikirim",
    ]),
    ("dent", [
        "dented", "dent ", "dents", "penyok", "penyokan",
        "bekas benturan", "kepenyok", "slightly bent", "bent ",
        "slightly bent",
    ]),
    ("scratch", [
        "scratched", "scratch", "scratches", "goresan", "gores",
        "lecet", "tergores",
    ]),
    ("peeling", [
        "paint peeling", "peeling", "peeled", "mengelupas",
        "cat mengelupas", "kulit mengelupas", "peel off",
    ]),
    ("leak", [
        "leaking", "leak ", "leaks", "leaky", "bocor", "merembes",
    ]),
    ("damage", [
        "damaged", "damage", "broken", "rusak", "pecah",
        "cracked", "crack", "patah", "sobek", "tear",
        "shattered", "snapped", "retak", "hancur",
        "broke during", "broken during", "lid broke", "lids broke",
        "one of the lids broke",
    ]),
    ("dirty", [
        "dirty", "kotor", "stained", "stain", "noda",
        "berdebu", "bernoda",
    ]),
    ("wrong_item", [
        "wrong item", "wrong size", "wrong color", "wrong model",
        "wrong bolt", "bolt is wrong", "one bolt is wrong",
        "salah barang", "salah ukuran", "salah warna", "salah model",
        "incorrect item", "wrong product", "incorrect",
        "received the wrong", "kirim salah",
        "ordered pink, received", "ordered white, received",
    ]),
    ("defective", [
        "defective", "cacat", "faulty", "malfunction",
        "not functioning", "fungsi rusak", "tidak presisi",
    ]),
    ("poor_quality", [
        "poor quality", "bad quality", "low quality",
        "kualitas buruk", "kualitas jelek", "kualitas kurang",
        "very thin", "tipis sekali", "thin material",
        "material tipis", "material murahan", "murahan",
        "cheap material", "flimsy", "ringkih",
        "thickness: thin", "material: thin", "design: poor",
        "mall-like", "low-quality", "not sharp", "disposable",
        "thin like zinc", "the pan is thin",
        "foams up", "foam up", "it's a shame", "a shame that",
    ]),
    ("disappointed", [
        "disappointed", "kecewa", "very disappointed",
        "sangat kecewa", "amat kecewa",
    ]),
    ("packaging", [
        "bad packaging", "poor packaging", "packaging damaged",
        "kemasan rusak", "kemasan buruk", "kardus penyok",
        "outer box damaged", "box damaged", "box dented",
    ]),
]

# Part keywords (single token usually). Order influences priority.
PART_RULES: List[Tuple[str, List[str]]] = [
    ("glass_lid",  ["lid glass", "glass lid", "tutup kaca", "tutup glass"]),
    ("lid",        ["lid", "tutup", " cover ", "cover.", "the cover"]),
    ("box",        ["bubble wrap", "outer box", "outer packaging",
                    "kardus", "cardboard", "box ", "box.", " box,",
                    "kemasan", "packaging"]),
    ("handle",     ["handle", "pegangan", "gagang"]),
    ("bolt",       ["bolt", "bolts", "baut"]),
    ("screw",      ["screw", "screws", "sekrup"]),
    ("knob",       ["knob", "knobs", "knop", "pemutar"]),
    ("cable",      ["cable", "charger", "kabel", "wire", "chord", "cord"]),
    ("steamer",    ["steamer", "kukusan", "pengukus"]),
    ("spatula",    ["spatula", "sodet"]),
    ("stove",      ["stove", "kompor"]),
    ("pan_body",   ["wokpan", "wok pan", "pan ", "pan,", "pan.",
                    "the pan", "wok", "panci", "pot ", "the pot",
                    "cookware", "kuali"]),
    ("design",     ["design", "desain", "model design"]),
    ("color",      ["color", "colour", "warna"]),
    ("material",   ["material", "bahan"]),
    ("appearance", ["appearance", "penampilan", "look", "tampilan"]),
    ("coating",    ["paint", "coating", "enamel", "cat ", " cat."]),
    ("size",       ["size", "ukuran", " dimensi"]),
]

NEGATIVE_PHRASES = [
    "tidak bagus", "kurang bagus", "not good", "not great", "not satisfied",
    "not happy", "not recommended", "belum bagus", "ga bagus", "gak bagus",
    "nggak bagus", "tidak puas", "not ok", "not okay", "no good",
    "very bad", "so bad", "really bad", "sangat buruk", "buruk sekali",
    "poor", "bad", "jelek", "buruk", "terrible", "awful",
    "disappointed", "kecewa", "complaint", "complain", "keluhan",
    "frustrated", "frustrasi", "marah", "menyedihkan", "menyebalkan",
    "it's a shame", "a shame that", "how can you", "not embarrassed",
    "i was shocked", "was a bit shocked", "takes so long",
]

POSITIVE_TOKENS = [
    "good", "bagus", "great", "excellent", "perfect", "nice", "love", "loved",
    "satisfied", "happy", "recommended", "amazing", "awesome",
    "puas", "memuaskan", "mantap", "keren", "suka",
]

NEGATION_RE = re.compile(
    r"\b(tidak|bukan|belum|jangan|tanpa|ga\b|gak\b|nggak|no\b|not\b|without|never|hardly|barely)\b",
    re.IGNORECASE,
)

STRUCTURED_PART_RE = re.compile(
    r"^\s*(appearance|quality|design|color|colour|material|size|packaging|"
    r"thickness|function|use|durability|power)\s*:",
    re.IGNORECASE | re.MULTILINE,
)

STRUCTURED_LINE_RE = re.compile(
    r"(?m)^\s*(design|size|material|thickness|function|use|durability|power|"
    r"appearance|quality|packaging|color|colour)\s*:\s*(.+?)\s*$",
    re.IGNORECASE,
)

ORDERED_RECEIVED_RE = re.compile(
    r"ordered\s+[\w\-]+,?\s+received\s+[\w\-]+",
    re.IGNORECASE,
)

_STRUCTURED_ISSUE_HINTS: List[Tuple[str, List[str]]] = [
    ("poor_quality", [
        r"\b(poor|bad|jelek|buruk|low[\s-]?quality|mall[\s-]?like|inappropriate|"
        r"not sharp|disposable|average|so[\s-]?so|inconsistent)\b",
        r"\b(thin|too small|smaller|quite small|not worth|thin like)\b",
    ]),
    ("not_as_described", [
        r"\b(wrong size|wrong color|do not match|does not match|not match)\b",
        r"\b(inappropriate|inconsistent information)\b",
    ]),
    ("damage", [r"\b(broke|broken|cracked)\b"]),
    ("dent", [r"\b(bent|penyok)\b"]),
    ("missing", [r"\b(not included|missing|wasn't included)\b"]),
]


def _phrase_pattern(phrase: str) -> re.Pattern:
    p = phrase.strip().lower()
    if any(ch in p for ch in " .,'-_!?") or len(p) <= 2:
        return re.compile(re.escape(p), re.IGNORECASE)
    return re.compile(r"\b" + re.escape(p) + r"\b", re.IGNORECASE)


def _compiled_rules(rules: List[Tuple[str, List[str]]]) -> List[Tuple[str, re.Pattern]]:
    out: List[Tuple[str, re.Pattern]] = []
    for key, phrases in rules:
        for phrase in sorted(phrases, key=len, reverse=True):
            out.append((key, _phrase_pattern(phrase)))
    return out


_ISSUE_COMPILED = _compiled_rules(ISSUE_RULES)
_PART_COMPILED = _compiled_rules(PART_RULES)
_NEG_COMPILED = [_phrase_pattern(p) for p in sorted(NEGATIVE_PHRASES, key=len, reverse=True)]
_POS_COMPILED = [_phrase_pattern(p) for p in sorted(POSITIVE_TOKENS, key=len, reverse=True)]
_MACRO_COMPILED: List[Tuple[re.Pattern, List[str], List[str]]] = [
    (re.compile(r"\b" + re.escape(tag) + r"\b", re.IGNORECASE), issues, parts)
    for tag, issues, parts in sorted(MACRO_RULES, key=lambda x: -len(x[0]))
]

# Catalog of all issue / part keys (stable display order)
ISSUE_ORDER = [
    "damage", "missing", "dent", "scratch", "not_fit", "not_working",
    "not_as_described", "spillage", "leak", "peeling", "wrong_item",
    "defective", "poor_quality", "packaging", "dirty", "disappointed",
    "change_of_mind",
]
PART_ORDER = [
    "lid", "glass_lid", "handle", "knob", "bolt", "screw",
    "pan_body", "box", "cable", "steamer", "spatula", "stove",
    "design", "color", "material", "appearance", "coating", "size",
]


def _is_negated(text: str, start: int) -> bool:
    window = text[max(0, start - 35):start]
    if not NEGATION_RE.search(window):
        return False
    matches = list(NEGATION_RE.finditer(window))
    last_neg = matches[-1].start() + max(0, start - 35)
    return (start - last_neg) <= 30


def _scan(text: str, compiled: List[Tuple[str, re.Pattern]], limit: int = 3,
          check_negation: bool = False) -> List[str]:
    found: List[str] = []
    seen: set = set()
    for key, pattern in compiled:
        if key in seen:
            continue
        m = pattern.search(text)
        if not m:
            continue
        if check_negation and _is_negated(text, m.start()):
            continue
        found.append(key)
        seen.add(key)
        if len(found) >= limit:
            break
    return found


def _structured_part_hints(text: str) -> List[str]:
    hints: List[str] = []
    for m in STRUCTURED_PART_RE.finditer(text):
        label = m.group(1).lower()
        if label in ("colour",):
            label = "color"
        if label == "quality":
            hints.append("material")
        elif label in ("appearance", "packaging"):
            hints.append("box" if label == "packaging" else "appearance")
        elif label in ("thickness", "function", "use", "durability", "power"):
            hints.append("material")
        else:
            hints.append(label)
    return hints


def _structured_issue_hints(text: str) -> List[str]:
    """Parse 'Design: Poor' / 'Thickness: thin' style lines into issue tags."""
    hints: List[str] = []
    seen: set = set()
    for m in STRUCTURED_LINE_RE.finditer(text):
        value = m.group(2).lower()
        for issue_key, patterns in _STRUCTURED_ISSUE_HINTS:
            if issue_key in seen:
                continue
            for pat in patterns:
                if re.search(pat, value, re.IGNORECASE):
                    hints.append(issue_key)
                    seen.add(issue_key)
                    break
    return hints


def _apply_macros(text: str) -> Tuple[List[str], List[str]]:
    """Match uppercase/underscore macro tags. Returns issues, parts (deduped)."""
    issues: List[str] = []
    parts: List[str] = []
    seen_i: set = set()
    seen_p: set = set()
    for pattern, m_issues, m_parts in _MACRO_COMPILED:
        if pattern.search(text):
            for i in m_issues:
                if i not in seen_i:
                    issues.append(i)
                    seen_i.add(i)
            for p in m_parts:
                if p not in seen_p:
                    parts.append(p)
                    seen_p.add(p)
    return issues, parts


def analyze_detail(detail: Any) -> Dict[str, Any]:
    raw = "" if detail is None or (isinstance(detail, float) and pd.isna(detail)) else str(detail)
    text = raw.strip()
    lower = text.lower()

    # 1) Macro tags first (they are explicit, no negation)
    macro_issues, macro_parts = _apply_macros(text)

    # 2) Structured field lines (Design: Poor, Thickness: thin, …)
    structured_issues = _structured_issue_hints(text)

    # 3) Ordered X, received Y
    pattern_issues: List[str] = []
    if ORDERED_RECEIVED_RE.search(lower):
        pattern_issues.append("wrong_item")

    # 4) Natural-language scan
    nlp_issues = _scan(lower, _ISSUE_COMPILED, limit=3)
    nlp_parts = _scan(lower, _PART_COMPILED, limit=3)

    # Merge with priority: macros → structured → patterns → NLP
    issues: List[str] = []
    parts: List[str] = []
    for i in macro_issues + structured_issues + pattern_issues + nlp_issues:
        if i not in issues and len(issues) < 3:
            issues.append(i)
    for p in macro_parts + nlp_parts:
        if p not in parts and len(parts) < 3:
            parts.append(p)

    # 5) Structured "Appearance: …" part hints if still room
    for hint in _structured_part_hints(text):
        if hint not in parts and len(parts) < 3:
            parts.append(hint)

    # 6) Sentiment (negation-aware on positive tokens only)
    sentiment = "neutral"
    for pat in _NEG_COMPILED:
        if pat.search(lower):
            sentiment = "negative"
            break
    if sentiment != "negative":
        for pat in _POS_COMPILED:
            m = pat.search(lower)
            if m and not _is_negated(lower, m.start()):
                sentiment = "positive"
                break

    # Any detected issue implies negative tone
    if issues and sentiment != "positive":
        sentiment = "negative"

    # 7) Light fallback for negative complaints with parts but no issue tag
    if not issues:
        if sentiment == "negative":
            if any(k in lower for k in ("missing", "wasn't included", "not included", "kurang")):
                issues.append("missing")
            elif re.search(r"\bbut\b.{0,80}\b(thin|small|shame|foam|wrong|fit|broke|missing)\b", lower):
                issues.append("poor_quality")
            elif parts and any(p for p in parts if p):
                issues.append("poor_quality")
            elif re.search(r"\bmatch|sesuai|wrong|salah\b", lower):
                issues.append("not_as_described")
            else:
                issues.append("disappointed")

    while len(issues) < 3:
        issues.append("")
    while len(parts) < 3:
        parts.append("")

    return {
        "issues": issues[:3],
        "parts": parts[:3],
        "sentiment": sentiment,
        "detail": text,
    }


MIN_SKU_LENGTH = 10
# Bundle delimiters: +  /  -  \  ,  and whitespace between tokens
SKU_SPLIT_PATTERN = re.compile(r"\s*[\+\/\\,\-]\s*|\s+")


def _valid_sku(token: str) -> bool:
    sku = token.strip()
    return bool(sku) and len(sku) >= MIN_SKU_LENGTH


def _sku_sort_key(sku: str) -> Tuple[int, str]:
    """FR-prefixed SKUs first, then alphabetical."""
    return (0 if sku.upper().startswith("FR") else 1, sku)


def split_skus(raw_sku: Any) -> List[str]:
    if raw_sku is None or (isinstance(raw_sku, float) and pd.isna(raw_sku)):
        return []
    text = str(raw_sku).strip()
    if not text or text.lower() in ("nan", "none"):
        return []
    parts = SKU_SPLIT_PATTERN.split(text)
    out: List[str] = []
    seen: set = set()
    for p in parts:
        sku = p.strip()
        if not _valid_sku(sku) or sku in seen:
            continue
        seen.add(sku)
        out.append(sku)
    if not out and _valid_sku(text):
        return [text.strip()]
    return sorted(out, key=_sku_sort_key)


def expand_rows(df: pd.DataFrame, col_map: Dict[str, str]) -> pd.DataFrame:
    records: List[Dict[str, Any]] = []
    for src_idx, row in df.iterrows():
        sku_raw = row.get(col_map["sku"])
        skus = split_skus(sku_raw)
        if not skus:
            continue
        detail = row.get(col_map.get("detail", ""), "")
        analysis = analyze_detail(detail)
        order_id = (
            str(row.get(col_map.get("order_id", ""), "") or "").strip()
            or str(row.get(col_map.get("order_serial_number", ""), "") or "").strip()
            or str(row.get(col_map.get("after_sales_order_no", ""), "") or "").strip()
            or f"row_{src_idx}"
        )
        base = {
            "source_row": int(src_idx),
            "order_key": order_id,
            "shop": str(row.get(col_map.get("shop", ""), "") or "").strip(),
            "after_sales_order_no": str(row.get(col_map.get("after_sales_order_no", ""), "") or "").strip(),
            "order_serial_number": str(row.get(col_map.get("order_serial_number", ""), "") or "").strip(),
            "order_id": str(row.get(col_map.get("order_id", ""), "") or "").strip(),
            "business_type": str(row.get(col_map.get("business_type", ""), "") or "").strip(),
            "after_sales_type": str(row.get(col_map.get("after_sales_type", ""), "") or "").strip(),
            "original_sku": str(sku_raw or "").strip(),
            "detail": analysis["detail"],
            "sentiment": analysis["sentiment"],
            "issue_1": analysis["issues"][0],
            "issue_2": analysis["issues"][1],
            "issue_3": analysis["issues"][2],
            "part_1": analysis["parts"][0],
            "part_2": analysis["parts"][1],
            "part_3": analysis["parts"][2],
            "categorized": bool(analysis["issues"][0]),
        }
        for sku in skus:
            if _valid_sku(sku):
                records.append({**base, "sku": sku})
    return pd.DataFrame(records)


def _order_level_breakdown(
    expanded: pd.DataFrame,
    field: str,
) -> List[Dict[str, Any]]:
    """Count unique orders per Business type / After-sales type (cols E & F)."""
    order_values: Dict[str, str] = {}
    for _, row in expanded.iterrows():
        order = str(row["order_key"])
        if order in order_values:
            continue
        val = str(row.get(field) or "").strip() or "Unknown"
        order_values[order] = val
    counter = Counter(order_values.values())
    total = len(order_values) or 1
    return [
        {"name": k, "count": int(v), "pct": _round_pct(v, total)}
        for k, v in counter.most_common()
    ]


def _round_pct(num: float, denom: float, ndigits: int = 1) -> float:
    if not denom:
        return 0.0
    return round((num / denom) * 100.0, ndigits)


def _top_issue_keys(counts: Counter, limit: int = 3) -> List[str]:
    return [k for k, v in counts.most_common(limit) if v > 0]


def resolve_sku_photos(skus: List[str]) -> Dict[str, Dict[str, Any]]:
    """Brand Material → SKU_Info link → Product Performance (same as Price Checker)."""
    from database import SessionLocal
    from models import FreemirName
    from services.price_checker_logic import _is_image_url, resolve_photo_maps_for_skus

    keys = sorted({(s or "").strip().upper() for s in skus if s and _valid_sku(s)})
    if not keys:
        return {}

    db = SessionLocal()
    try:
        brand_meta, photo_map = resolve_photo_maps_for_skus(db, set(keys))
        link_by_sku: Dict[str, str] = {}
        for row in db.query(FreemirName).filter(FreemirName.sku.in_(keys)).all():
            link_by_sku[(row.sku or "").strip().upper()] = (row.link or "").strip()

        out: Dict[str, Dict[str, Any]] = {}
        for sku in keys:
            meta = brand_meta.get(sku, {})
            link = link_by_sku.get(sku, "")
            brand_id = meta.get("materialId")
            brand_url = (meta.get("url") or "").strip()
            preview = (meta.get("previewUrl") or "").strip()
            image = None
            source = None
            if brand_id:
                image = preview or brand_url or None
                source = "brand_material"
            elif _is_image_url(link):
                image = link
                source = "sku_info"
            else:
                pp = photo_map.get(sku) or photo_map.get(sku.upper()) or photo_map.get(sku.lower())
                if pp:
                    image = pp
                    source = "product_performance"
            out[sku] = {
                "image": image,
                "imageSource": source,
                "brandMaterialId": brand_id,
                "previewUrl": preview or None,
                "previewGcsPath": meta.get("previewGcsPath") or "",
            }
        return out
    finally:
        db.close()


def _pct_to_excel(pct_display: float) -> float:
    """Convert display percent (e.g. 89.3) to Excel ratio (0.893)."""
    return round((pct_display or 0) / 100.0, 4)


def _issue_label(key: str) -> str:
    return key.replace("_", " ").title()


def _build_store_matrix(
    expanded: pd.DataFrame,
    ordered_issue_keys: List[str],
) -> Dict[str, Any]:
    """Per-store order counts and issue breakdown (order-level dedup)."""
    store_orders: Dict[str, set] = defaultdict(set)
    store_problem_orders: Dict[str, set] = defaultdict(set)
    store_issue_orders: Dict[str, Counter] = defaultdict(Counter)

    order_issues: Dict[Tuple[str, str], set] = defaultdict(set)
    for _, row in expanded.iterrows():
        shop = str(row.get("shop") or "").strip() or "Unknown"
        order = str(row["order_key"])
        store_orders[shop].add(order)
        if row.get("categorized"):
            store_problem_orders[shop].add(order)
        for col in ("issue_1", "issue_2", "issue_3"):
            iss = str(row.get(col, "") or "").strip()
            if iss:
                order_issues[(shop, order)].add(iss)

    for (shop, _order), issues in order_issues.items():
        for iss in issues:
            store_issue_orders[shop][iss] += 1

    rows: List[Dict[str, Any]] = []
    for shop in sorted(
        store_orders.keys(),
        key=lambda s: (-len(store_problem_orders[s]), s),
    ):
        total = len(store_orders[shop])
        problem = len(store_problem_orders[shop])
        row: Dict[str, Any] = {
            "store": shop,
            "total_orders": total,
            "problem_orders": problem,
            "problem_pct": min(100.0, _round_pct(problem, total)),
        }
        store_counts = store_issue_orders[shop]
        for k in ordered_issue_keys:
            cnt = int(store_counts.get(k, 0))
            row[k] = cnt
            row[f"{k}_pct"] = min(100.0, _round_pct(cnt, total))
        row["highlight_issues"] = _top_issue_keys(store_counts, 3)
        rows.append(row)

    return {
        "columns": ["store", "total_orders", "problem_orders"] + ordered_issue_keys,
        "issue_keys": ordered_issue_keys,
        "rows": rows,
    }


def _build_sentiment_by_mention(
    expanded: pd.DataFrame,
    tag_columns: Tuple[str, ...],
    preferred_order: List[str],
) -> List[Dict[str, Any]]:
    """
    Sentiment split per part/issue tag (one count per tag per expanded row, deduped).
    Example: material negative vs design positive on different rows.
    """
    counts: Dict[str, Counter] = defaultdict(Counter)
    for _, row in expanded.iterrows():
        sentiment = str(row.get("sentiment") or "neutral").strip().lower()
        if sentiment not in ("negative", "positive", "neutral"):
            sentiment = "neutral"
        seen: set = set()
        for col in tag_columns:
            tag = str(row.get(col) or "").strip()
            if not tag or tag in seen:
                continue
            seen.add(tag)
            counts[tag][sentiment] += 1

    if not counts:
        return []

    ordered_tags = [k for k in preferred_order if k in counts]
    remaining = sorted(
        (k for k in counts if k not in ordered_tags),
        key=lambda k: (-sum(counts[k].values()), k),
    )
    ordered_tags.extend(remaining)

    rows: List[Dict[str, Any]] = []
    for tag in ordered_tags:
        c = counts[tag]
        total = sum(c.values()) or 1
        rows.append({
            "tag": tag,
            "mentions": int(total),
            "negative": int(c.get("negative", 0)),
            "positive": int(c.get("positive", 0)),
            "neutral": int(c.get("neutral", 0)),
            "negative_pct": _round_pct(c.get("negative", 0), total),
            "positive_pct": _round_pct(c.get("positive", 0), total),
            "neutral_pct": _round_pct(c.get("neutral", 0), total),
        })
    return rows


def build_summaries(expanded: pd.DataFrame, source_rows: int) -> Dict[str, Any]:
    empty = {
        "stats": {
            "source_rows": source_rows,
            "expanded_rows": 0,
            "unique_skus": 0,
            "unique_orders": 0,
            "bundle_splits": 0,
            "categorized_rows": 0,
            "uncategorized_rows": 0,
            "categorized_pct": 0.0,
            "uncategorized_pct": 0.0,
            "sentiment_counts": {"negative": 0, "positive": 0, "neutral": 0},
            "sentiment_pct": {"negative": 0.0, "positive": 0.0, "neutral": 0.0},
        },
        "top_issues": [],
        "top_parts": [],
        "issue_totals": {},
        "sku_matrix": {"columns": [], "rows": []},
        "store_matrix": {"columns": [], "rows": []},
        "business_types": [],
        "after_sales_types": [],
        "sentiment_by_mention": {"parts": [], "issues": []},
    }
    if expanded.empty:
        return empty

    issue_counter: Counter = Counter()
    part_counter: Counter = Counter()
    # SKU matrix: orders = expanded rows per SKU (matches Expanded Data count per SKU)
    sku_row_count: Dict[str, int] = defaultdict(int)
    sku_issue_rows: Dict[str, Counter] = defaultdict(Counter)
    sku_categorized_rows: Dict[str, int] = defaultdict(int)

    for _, row in expanded.iterrows():
        sku = row["sku"]
        sku_row_count[sku] += 1
        if row.get("categorized"):
            sku_categorized_rows[sku] += 1
        row_issues: set = set()
        for col in ("issue_1", "issue_2", "issue_3"):
            v = str(row.get(col, "") or "").strip()
            if v:
                issue_counter[v] += 1
                row_issues.add(v)
        for iss in row_issues:
            sku_issue_rows[sku][iss] += 1
        for col in ("part_1", "part_2", "part_3"):
            v = str(row.get(col, "") or "").strip()
            if v:
                part_counter[v] += 1

    total_issue_mentions = sum(issue_counter.values())
    total_part_mentions = sum(part_counter.values())

    # Order issue columns: stable preferred order, then any unseen by frequency
    ordered_issue_keys = [k for k in ISSUE_ORDER if k in issue_counter] + \
                         [k for k, _ in issue_counter.most_common() if k not in ISSUE_ORDER]

    matrix_rows: List[Dict[str, Any]] = []
    eligible_skus = [s for s in sku_row_count if _valid_sku(s)]
    for sku in sorted(
        eligible_skus,
        key=lambda s: (-sku_row_count[s], _sku_sort_key(s), s),
    ):
        orders = sku_row_count[sku]
        counts = sku_issue_rows[sku]
        mentions = sum(counts.values())
        categorized_n = sku_categorized_rows[sku]
        row: Dict[str, Any] = {
            "sku": sku,
            "orders": orders,
            "mentions": mentions,
            "categorized_orders": categorized_n,
            "categorized_pct": _round_pct(categorized_n, orders),
            "highlight_issues": _top_issue_keys(counts, 3),
        }
        for k in ordered_issue_keys:
            cnt = int(counts.get(k, 0))
            row[k] = cnt
            row[f"{k}_pct"] = min(100.0, _round_pct(cnt, orders))
        matrix_rows.append(row)

    categorized = int(expanded["categorized"].sum())
    uncategorized = len(expanded) - categorized
    sentiments = expanded["sentiment"].value_counts().to_dict()
    neg = int(sentiments.get("negative", 0))
    pos = int(sentiments.get("positive", 0))
    neu = int(sentiments.get("neutral", 0))
    total_sent = neg + pos + neu or 1
    bundle_splits = int(
        expanded["original_sku"].astype(str).str.contains(
            r"[\+\/\\,\-]|\s{2,}", regex=True, na=False,
        ).sum()
    )
    business_types = _order_level_breakdown(expanded, "business_type")
    after_sales_types = _order_level_breakdown(expanded, "after_sales_type")

    return {
        "stats": {
            "source_rows": source_rows,
            "expanded_rows": len(expanded),
            "unique_skus": int(
                expanded.loc[expanded["sku"].astype(str).str.len() >= MIN_SKU_LENGTH, "sku"].nunique()
            ),
            "unique_orders": expanded["order_key"].nunique(),
            "bundle_splits": bundle_splits,
            "categorized_rows": categorized,
            "uncategorized_rows": uncategorized,
            "categorized_pct": _round_pct(categorized, len(expanded)),
            "uncategorized_pct": _round_pct(uncategorized, len(expanded)),
            "sentiment_counts": {"negative": neg, "positive": pos, "neutral": neu},
            "sentiment_pct": {
                "negative": _round_pct(neg, total_sent),
                "positive": _round_pct(pos, total_sent),
                "neutral":  _round_pct(neu, total_sent),
            },
        },
        "top_issues": [
            {"name": k, "count": int(v), "pct": _round_pct(v, total_issue_mentions)}
            for k, v in issue_counter.most_common(20)
        ],
        "top_parts": [
            {"name": k, "count": int(v), "pct": _round_pct(v, total_part_mentions)}
            for k, v in part_counter.most_common(20)
        ],
        "issue_totals": {k: int(v) for k, v in issue_counter.items()},
        "sku_matrix": {
            "columns": ["sku", "orders", "mentions"] + ordered_issue_keys,
            "issue_keys": ordered_issue_keys,
            "rows": matrix_rows,
        },
        "store_matrix": _build_store_matrix(expanded, ordered_issue_keys),
        "business_types": business_types,
        "after_sales_types": after_sales_types,
        "sentiment_by_mention": {
            "parts": _build_sentiment_by_mention(
                expanded, ("part_1", "part_2", "part_3"), PART_ORDER,
            ),
            "issues": _build_sentiment_by_mention(
                expanded, ("issue_1", "issue_2", "issue_3"), ISSUE_ORDER,
            ),
        },
    }


def process_sku_review(content: bytes, filename: str) -> Tuple[pd.DataFrame, Dict[str, Any], Dict[str, str]]:
    raw_df = _read_upload(content, filename)
    raw_df = raw_df.dropna(how="all")
    col_map = _map_columns(raw_df)
    source_rows = len(raw_df)
    expanded = expand_rows(raw_df, col_map)
    summaries = build_summaries(expanded, source_rows)
    return expanded, summaries, col_map


# ---------------------------------------------------------------------------
# Excel export — borders, percent format, Store Matrix
# ---------------------------------------------------------------------------
# Navy headers (white text) + yellow top-3 data cells
COLOR_NAVY = "1E3A5F"
COLOR_NAVY_MID = "334155"
COLOR_TOP3_CELL = "FEF9C3"
HEADER_FONT_COLOR = "FFFFFF"


def _header_font_navy():
    from openpyxl.styles import Font

    return Font(bold=True, color=HEADER_FONT_COLOR, size=11)


def _matrix_export_df(
    rows: List[Dict[str, Any]],
    issue_keys: List[str],
    id_col: str,
    id_header: str,
    meta_cols: List[Tuple[str, str]],
    *,
    include_photo_col: bool = False,
) -> pd.DataFrame:
    """ID [+ Photo] + meta | all issue integers | all issue percents (grouped)."""
    out_rows: List[Dict[str, Any]] = []
    for r in rows:
        row: Dict[str, Any] = {id_header: r.get(id_col, "")}
        if include_photo_col:
            row["Photo"] = ""
        for key, hdr in meta_cols:
            if key.endswith("_pct") or hdr.endswith("%"):
                row[hdr] = _pct_to_excel(r.get(key, 0))
            else:
                row[hdr] = r.get(key, 0)
        for k in issue_keys:
            row[f"[Count] {_issue_label(k)}"] = int(r.get(k, 0) or 0)
        for k in issue_keys:
            row[f"[%] {_issue_label(k)}"] = _pct_to_excel(r.get(f"{k}_pct", 0))
        out_rows.append(row)
    return pd.DataFrame(out_rows)


def _append_category_table(
    summary_rows: List[List[Any]],
    title: str,
    items: List[Dict[str, Any]],
) -> None:
    summary_rows.append([])
    summary_rows.append([title, "Count", "Percent"])
    for item in items:
        summary_rows.append([
            item["name"], item["count"], _pct_to_excel(item["pct"]),
        ])


def _sentiment_mention_export_df(summaries: Dict[str, Any]) -> pd.DataFrame:
    """Flat table for Excel: part + issue rows with sentiment counts and %."""
    records: List[Dict[str, Any]] = []
    block = summaries.get("sentiment_by_mention") or {}
    for kind, key in (("Part", "parts"), ("Issue", "issues")):
        for item in block.get(key) or []:
            records.append({
                "Type": kind,
                "Tag": _issue_label(item.get("tag", "")),
                "Mentions": item.get("mentions", 0),
                "Negative": item.get("negative", 0),
                "Positive": item.get("positive", 0),
                "Neutral": item.get("neutral", 0),
                "Negative %": _pct_to_excel(item.get("negative_pct", 0)),
                "Positive %": _pct_to_excel(item.get("positive_pct", 0)),
                "Neutral %": _pct_to_excel(item.get("neutral_pct", 0)),
            })
    return pd.DataFrame(records)


def _ai_insights_sheet_rows(ai: Dict[str, Any]) -> List[List[Any]]:
    """Two-column layout for dedicated AI Insights worksheet."""
    rows: List[List[Any]] = [["Section", "Content"]]
    rows.append(["Executive summary", ai.get("executive_summary", "")])
    rows.append(["", ""])
    rows.append(["Key findings", ""])
    for line in ai.get("key_findings") or []:
        rows.append(["", f"• {line}"])
    rows.append(["", ""])
    rows.append(["Recommendations", ""])
    for line in ai.get("recommendations") or []:
        rows.append(["", f"• {line}"])
    if ai.get("priority_skus"):
        rows.append(["", ""])
        rows.append(["Priority SKUs", ", ".join(ai.get("priority_skus", []))])
    if ai.get("priority_stores"):
        rows.append(["Priority stores", ", ".join(ai.get("priority_stores", []))])
    if ai.get("model"):
        rows.append(["", ""])
        rows.append(["Model", ai.get("model", "")])
    if ai.get("disclaimer"):
        rows.append(["Note", ai.get("disclaimer", "")])
    return rows


def export_sku_review_excel(
    expanded: pd.DataFrame,
    summaries: Dict[str, Any],
    *,
    include_photos: bool = False,
    photo_by_sku: Optional[Dict[str, Dict[str, Any]]] = None,
    ai_insights: Optional[Dict[str, Any]] = None,
) -> bytes:
    stats = summaries.get("stats", {})
    sentiment_counts = stats.get("sentiment_counts", {})
    sentiment_pct = stats.get("sentiment_pct", {})
    issue_keys = summaries.get("sku_matrix", {}).get("issue_keys", [])

    # --- Summary sheet (structured blocks) ---
    summary_rows: List[List[Any]] = []
    summary_rows.append(["Metric", "Value"])
    summary_rows.append(["Source rows", stats.get("source_rows", 0)])
    summary_rows.append(["Expanded SKU rows", stats.get("expanded_rows", 0)])
    summary_rows.append(["Unique SKUs", stats.get("unique_skus", 0)])
    summary_rows.append(["Unique orders", stats.get("unique_orders", 0)])
    summary_rows.append(["Bundle split rows", stats.get("bundle_splits", 0)])
    summary_rows.append(["Categorized rows", stats.get("categorized_rows", 0)])
    summary_rows.append(["Uncategorized rows", stats.get("uncategorized_rows", 0)])
    summary_rows.append(["Categorized %", _pct_to_excel(stats.get("categorized_pct", 0))])
    summary_rows.append(["Uncategorized %", _pct_to_excel(stats.get("uncategorized_pct", 0))])
    summary_rows.append([])
    summary_rows.append(["Sentiment", "Count", "Percent"])
    summary_rows.append([
        "Negative", sentiment_counts.get("negative", 0),
        _pct_to_excel(sentiment_pct.get("negative", 0)),
    ])
    summary_rows.append([
        "Positive", sentiment_counts.get("positive", 0),
        _pct_to_excel(sentiment_pct.get("positive", 0)),
    ])
    summary_rows.append([
        "Neutral", sentiment_counts.get("neutral", 0),
        _pct_to_excel(sentiment_pct.get("neutral", 0)),
    ])
    summary_rows.append([])
    summary_rows.append(["Top issues", "Count", "Percent of mentions"])
    for item in summaries.get("top_issues", []):
        summary_rows.append([
            _issue_label(item["name"]), item["count"], _pct_to_excel(item["pct"]),
        ])
    summary_rows.append([])
    summary_rows.append(["Top parts / details", "Count", "Percent of mentions"])
    for item in summaries.get("top_parts", []):
        summary_rows.append([
            _issue_label(item["name"]), item["count"], _pct_to_excel(item["pct"]),
        ])
    _append_category_table(
        summary_rows, "Business type (col E)", summaries.get("business_types", []),
    )
    _append_category_table(
        summary_rows, "After-sales type (col F)", summaries.get("after_sales_types", []),
    )
    sku_matrix_rows = summaries.get("sku_matrix", {}).get("rows", [])
    sku_df = _matrix_export_df(
        sku_matrix_rows,
        issue_keys,
        "sku",
        "SKU",
        [
            ("orders", "Orders"),
            ("mentions", "Mentions"),
        ],
        include_photo_col=include_photos,
    )
    store_df = _matrix_export_df(
        summaries.get("store_matrix", {}).get("rows", []),
        issue_keys,
        "store",
        "Store",
        [
            ("total_orders", "Total Orders"),
            ("problem_orders", "Problem Orders"),
            ("problem_pct", "Problem %"),
        ],
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(
            writer, sheet_name="Summary", index=False, header=False,
        )
        if not sku_df.empty:
            sku_df.to_excel(writer, sheet_name="SKU Matrix", index=False)
        if not store_df.empty:
            store_df.to_excel(writer, sheet_name="Store Matrix", index=False)
        sent_df = _sentiment_mention_export_df(summaries)
        if not sent_df.empty:
            sent_df.to_excel(writer, sheet_name="Sentiment by Mention", index=False)
        if not expanded.empty:
            export_cols = [
                "shop", "after_sales_order_no", "order_serial_number", "order_id",
                "business_type", "after_sales_type", "original_sku", "sku",
                "issue_1", "issue_2", "issue_3", "part_1", "part_2", "part_3",
                "sentiment", "categorized", "detail",
            ]
            cols = [c for c in export_cols if c in expanded.columns]
            exp_out = expanded[expanded["sku"].astype(str).str.len() >= MIN_SKU_LENGTH]
            exp_out[cols].to_excel(writer, sheet_name="Expanded Data", index=False)
        if ai_insights:
            pd.DataFrame(_ai_insights_sheet_rows(ai_insights)).to_excel(
                writer, sheet_name="AI Insights", index=False, header=False,
            )

    buf.seek(0)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(buf)
        thin = Side(style="thin", color="94A3B8")
        medium = Side(style="medium", color="475569")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        outer_border = Border(left=medium, right=medium, top=medium, bottom=medium)
        bold_white = _header_font_navy()
        bold_section = Font(bold=True, color=HEADER_FONT_COLOR, size=10)
        top3_fill = PatternFill("solid", fgColor=COLOR_TOP3_CELL)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        pct_fmt = "0.0%"

        def _apply_block_border(ws, r1: int, r2: int, c1: int, c2: int) -> None:
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = cell_border
            for c in range(c1, c2 + 1):
                ws.cell(row=r1, column=c).border = Border(
                    left=thin, right=thin, top=medium, bottom=thin,
                )
                ws.cell(row=r2, column=c).border = Border(
                    left=thin, right=thin, top=thin, bottom=medium,
                )
            for r in range(r1, r2 + 1):
                ws.cell(row=r, column=c1).border = Border(
                    left=medium, right=thin, top=thin, bottom=thin,
                )
                ws.cell(row=r, column=c2).border = Border(
                    left=thin, right=medium, top=thin, bottom=thin,
                )
            ws.cell(row=r1, column=c1).border = Border(
                left=medium, right=thin, top=medium, bottom=thin,
            )
            ws.cell(row=r1, column=c2).border = Border(
                left=thin, right=medium, top=medium, bottom=thin,
            )
            ws.cell(row=r2, column=c1).border = Border(
                left=medium, right=thin, top=thin, bottom=medium,
            )
            ws.cell(row=r2, column=c2).border = Border(
                left=thin, right=medium, top=thin, bottom=medium,
            )

        def _format_pct_columns(ws, header_row: int = 1) -> None:
            headers = [str(ws.cell(row=header_row, column=c).value or "") for c in range(1, ws.max_column + 1)]
            for c, hdr in enumerate(headers, start=1):
                if "%" in hdr or "Percent" in hdr:
                    for r in range(header_row + 1, ws.max_row + 1):
                        cell = ws.cell(row=r, column=c)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = pct_fmt
                            cell.alignment = center

        _TYPE_TABLE_TITLES = {
            "Business type (col E)", "After-sales type (col F)",
        }
        _ISSUE_TABLE_TITLES = {"Top issues", "Top parts / details"}

        def _style_summary_sheet(ws) -> None:
            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 12
            header_titles = {
                "Metric", "Sentiment", "Count", "Percent", "Value",
                "Percent of mentions",
            }
            navy_hdr = PatternFill("solid", fgColor=COLOR_NAVY)
            for r in range(1, ws.max_row + 1):
                label = ws.cell(row=r, column=1).value
                label_s = str(label) if label is not None else ""
                is_header_row = (
                    label_s in header_titles
                    or label_s in _TYPE_TABLE_TITLES
                    or label_s in _ISSUE_TABLE_TITLES
                )
                for c in range(1, min(ws.max_column, 3) + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = center
                    cell.border = cell_border
                    if is_header_row:
                        cell.fill = navy_hdr
                        cell.font = bold_white
            _format_pct_columns(ws, 1)
            for r in range(1, ws.max_row + 1):
                for c in range(1, 4):
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, float) and 0 <= v <= 1:
                        ws.cell(row=r, column=c).number_format = pct_fmt

        def _issue_key_from_hdr(hdr: str, issue_keys: List[str]) -> Optional[str]:
            prefix = "[Count] " if hdr.startswith("[Count] ") else "[%] " if hdr.startswith("[%] ") else ""
            if not prefix:
                return None
            label = hdr[len(prefix):]
            for k in issue_keys:
                if _issue_label(k) == label:
                    return k
            return None

        def _style_issue_matrix_sheet(
            ws,
            id_header: str,
            freeze_col: str,
            issue_keys: List[str],
            matrix_rows: List[Dict[str, Any]],
            id_field: str,
            *,
            bottom_border_only: bool = False,
        ) -> None:
            if ws.max_row < 2:
                return
            headers = [str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)]
            id_col_idx = headers.index(id_header) + 1 if id_header in headers else 1
            photo_col_idx = headers.index("Photo") + 1 if "Photo" in headers else None
            navy_hdr = PatternFill("solid", fgColor=COLOR_NAVY)
            highlight_by_id = {
                str(r.get(id_field, "")): set(r.get("highlight_issues") or [])
                for r in matrix_rows
            }

            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = navy_hdr
                cell.font = bold_white
                cell.alignment = center
                cell.border = Border(left=thin, right=thin, top=thin, bottom=medium)

            if bottom_border_only:
                row_sep = Border(bottom=thin)
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).border = row_sep
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=1, column=c).border = Border(
                        left=thin, right=thin, top=thin, bottom=medium,
                    )
            else:
                _apply_block_border(ws, 1, ws.max_row, 1, ws.max_column)

            ws.column_dimensions[get_column_letter(id_col_idx)].width = 36
            if photo_col_idx:
                ws.column_dimensions[get_column_letter(photo_col_idx)].width = 11
            for c in range(1, ws.max_column + 1):
                if c not in (id_col_idx, photo_col_idx):
                    ws.column_dimensions[get_column_letter(c)].width = 10

            for r in range(2, ws.max_row + 1):
                row_id = str(ws.cell(row=r, column=id_col_idx).value or "")
                highlights = highlight_by_id.get(row_id, set())
                ws.cell(row=r, column=id_col_idx).alignment = left_wrap
                for c in range(1, ws.max_column + 1):
                    if c == id_col_idx:
                        continue
                    ws.cell(row=r, column=c).alignment = center
                    hdr = headers[c - 1]
                    ik = _issue_key_from_hdr(hdr, issue_keys)
                    if ik and ik in highlights:
                        ws.cell(row=r, column=c).fill = top3_fill

            _format_pct_columns(ws, 1)
            ws.freeze_panes = f"{freeze_col}2"
            ws.row_dimensions[1].height = 32

        def _embed_sku_matrix_photos(ws) -> None:
            if not include_photos or not photo_by_sku:
                return
            headers = [str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)]
            if "Photo" not in headers or "SKU" not in headers:
                return
            photo_col = headers.index("Photo") + 1
            sku_col = headers.index("SKU") + 1
            from concurrent.futures import ThreadPoolExecutor, as_completed
            from openpyxl.drawing.image import Image as XLImage

            from services.excel_picture_utils import (
                FRAME_H_PX,
                FRAME_W_PX,
                fetch_framed_image_bytes,
            )

            img_cache: Dict[str, Optional[bytes]] = {}
            row_jobs: List[Tuple[int, str, str, str]] = []
            for r in range(2, ws.max_row + 1):
                sku_val = str(ws.cell(row=r, column=sku_col).value or "").strip()
                sku_key = sku_val.upper()
                pinfo = photo_by_sku.get(sku_key) or photo_by_sku.get(sku_val) or {}
                gcs_path = (pinfo.get("previewGcsPath") or "").strip()
                url = (pinfo.get("previewUrl") or pinfo.get("image") or "").strip()
                if gcs_path or url:
                    row_jobs.append((r, sku_key, url, gcs_path))

            def _load(job: Tuple[int, str, str, str]) -> Tuple[int, Optional[bytes]]:
                r, _sku, url, gcs_path = job
                framed = fetch_framed_image_bytes(
                    url, img_cache, gcs_object_path=gcs_path,
                )
                return r, framed

            loaded: Dict[int, bytes] = {}
            max_workers = min(16, max(1, len(row_jobs)))
            if row_jobs:
                with ThreadPoolExecutor(max_workers=max_workers) as pool:
                    futures = [pool.submit(_load, job) for job in row_jobs]
                    for fut in as_completed(futures):
                        try:
                            r, framed = fut.result()
                            if framed:
                                loaded[r] = framed
                        except Exception:
                            pass

            for r, framed in loaded.items():
                try:
                    img = XLImage(io.BytesIO(framed))
                    img.width = FRAME_W_PX
                    img.height = FRAME_H_PX
                    ws.row_dimensions[r].height = 44
                    ws.add_image(img, f"{get_column_letter(photo_col)}{r}")
                except Exception:
                    continue

        store_matrix_rows = summaries.get("store_matrix", {}).get("rows", [])

        if "Summary" in wb.sheetnames:
            _style_summary_sheet(wb["Summary"])

        if "SKU Matrix" in wb.sheetnames:
            sku_freeze = "C" if include_photos else "B"
            _style_issue_matrix_sheet(
                wb["SKU Matrix"],
                "SKU",
                sku_freeze,
                issue_keys,
                sku_matrix_rows,
                "sku",
            )
            _embed_sku_matrix_photos(wb["SKU Matrix"])

        if "Store Matrix" in wb.sheetnames:
            _style_issue_matrix_sheet(
                wb["Store Matrix"],
                "Store",
                "B",
                issue_keys,
                store_matrix_rows,
                "store",
                bottom_border_only=True,
            )

        if "Expanded Data" in wb.sheetnames:
            ws = wb["Expanded Data"]
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = PatternFill("solid", fgColor=COLOR_NAVY)
                cell.font = bold_white
                cell.alignment = center
                cell.border = cell_border
            _apply_block_border(ws, 1, ws.max_row, 1, ws.max_column)
            ws.column_dimensions["A"].width = 32
            ws.freeze_panes = "A2"

        if "Sentiment by Mention" in wb.sheetnames:
            ws = wb["Sentiment by Mention"]
            navy_hdr = PatternFill("solid", fgColor=COLOR_NAVY)
            widths = {"A": 10, "B": 22, "C": 12, "D": 11, "E": 11, "F": 11, "G": 12, "H": 12, "I": 12}
            for col_letter, w in widths.items():
                ws.column_dimensions[col_letter].width = w
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = navy_hdr
                cell.font = bold_white
                cell.alignment = center
                cell.border = cell_border
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = cell_border
                    cell.alignment = center if c >= 3 else left_wrap
            _format_pct_columns(ws, 1)
            ws.freeze_panes = "A2"

        if "AI Insights" in wb.sheetnames:
            ws = wb["AI Insights"]
            navy_hdr = PatternFill("solid", fgColor=COLOR_NAVY)
            section_fill = PatternFill("solid", fgColor="E2E8F0")
            section_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
            wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
            section_labels = {
                "Executive summary", "Key findings", "Recommendations",
                "Priority SKUs", "Priority stores", "Model", "Note",
            }
            for c in range(1, 3):
                cell = ws.cell(row=1, column=c)
                cell.fill = navy_hdr
                cell.font = bold_white
                cell.alignment = center
                cell.border = cell_border
            ws.column_dimensions["A"].width = 26
            ws.column_dimensions["B"].width = 100
            for r in range(2, ws.max_row + 1):
                label = str(ws.cell(row=r, column=1).value or "")
                is_section = label in section_labels
                for c in range(1, 3):
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = wrap_left
                    cell.border = cell_border
                    if is_section and c == 1:
                        cell.font = section_font
                        cell.fill = section_fill
                content = ws.cell(row=r, column=2).value
                if content:
                    ln = len(str(content))
                    if ln > 60:
                        ws.row_dimensions[r].height = min(120, max(32, ln // 3))
            ws.freeze_panes = "A2"

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()
    except Exception:
        import traceback
        print(traceback.format_exc())
        buf.seek(0)
        return buf.getvalue()
